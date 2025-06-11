from flask import Flask, render_template, request, redirect, url_for, send_from_directory, flash
import os
import uuid
import subprocess
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import tempfile
import shutil
import datetime
import json
import csv
from copy import copy

app = Flask(__name__)
app.secret_key = "your-secret-key"

UPLOAD_FOLDER = os.path.join("static", "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# === Utility functions merged from nta_utils.py ===

CONFIG_PATH = "config.json"
SETTINGS_PATH = "settings.json"

def load_config():
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, "r") as f:
            return json.load(f)
    return {}

def save_template_path(path):
    config = load_config()
    config["template_path"] = path
    with open(CONFIG_PATH, "w") as f:
        json.dump(config, f, indent=4)

def load_template_path():
    config = load_config()
    template_path = config.get("template_path")
    if not template_path or not os.path.exists(template_path):
        raise FileNotFoundError("Saved template path not found or does not exist.")
    return template_path

def save_settings(settings):
    with open(SETTINGS_PATH, "w") as f:
        json.dump(settings, f, indent=4)

def load_settings():
    if os.path.exists(SETTINGS_PATH):
        with open(SETTINGS_PATH, "r") as f:
            return json.load(f)
    return {}

# --- Processing helpers
def fill_plate(ws, assay_title, pseudotypes, sample_ids, num_pseudotypes):
    ws['B3'] = assay_title

    if num_pseudotypes == 1:
        val = pseudotypes[0]
        for cell in ['B3', 'E3', 'H3', 'K3']:
            ws[cell] = val
        sample_cells = ['B4', 'E4', 'H4', 'K4']
        for i_cell, cell in enumerate(sample_cells):
            ws[cell] = sample_ids[i_cell] if i_cell < len(sample_ids) else ''
    elif num_pseudotypes == 2:
        ws['B3'] = pseudotypes[0]
        ws['E3'] = pseudotypes[0]
        ws['H3'] = pseudotypes[1]
        ws['K3'] = pseudotypes[1]
        ws['B4'] = sample_ids[0] if len(sample_ids) > 0 else ''
        ws['H4'] = sample_ids[0] if len(sample_ids) > 0 else ''
        ws['E4'] = sample_ids[1] if len(sample_ids) > 1 else ''
        ws['K4'] = sample_ids[1] if len(sample_ids) > 1 else ''
    elif num_pseudotypes == 3:
        ws['B3'] = pseudotypes[0]
        ws['E3'] = pseudotypes[1]
        ws['H3'] = pseudotypes[2]
        ws['K3'] = ''
        val = sample_ids[0] if len(sample_ids) > 0 else ''
        ws['B4'] = val
        ws['E4'] = val
        ws['H4'] = val
        ws['K4'] = ''
    elif num_pseudotypes == 4:
        for idx, cell in enumerate(['B3', 'E3', 'H3', 'K3']):
            ws[cell] = pseudotypes[idx] if idx < len(pseudotypes) else ''
        val = sample_ids[0] if len(sample_ids) > 0 else ''
        for cell in ['B4', 'E4', 'H4', 'K4']:
            ws[cell] = val
    else:
        for cell in ['B3', 'E3', 'H3', 'K3', 'B4', 'E4', 'H4', 'K4']:
            ws[cell] = ''


def copy_worksheet(source_ws, target_ws):
    for row in source_ws.iter_rows():
        for cell in row:
            target_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.border = copy(cell.border)
                target_cell.fill = copy(cell.fill)
                target_cell.number_format = cell.number_format
                target_cell.protection = copy(cell.protection)
                target_cell.alignment = copy(cell.alignment)

def load_csv_blocks(csv_path):
    blocks = []
    current_block = []
    with open(csv_path, newline='') as f:
        reader = csv.reader(f)
        for row in reader:
            # If row is blank or all cells empty, treat as block separator
            if not any(cell.strip() for cell in row):
                if current_block:
                    blocks.append(current_block)
                    current_block = []
            else:
                current_block.append(row[:12])
        if current_block:
            blocks.append(current_block)
    return blocks

def add_default_to_final_titres(output_path):
    wb = load_workbook(output_path)
    summary = wb["Summary"]
    plate1 = wb["Plate 1"]
    a5_val = plate1["A5"].value

    try:
        a5_val = round(float(a5_val))
    except:
        a5_val = ""

    for row in summary.iter_rows(min_row=2, max_row=summary.max_row, min_col=4, max_col=10):
        for cell in row:
            if cell.value in (None, "") and a5_val:
                cell.value = f"<{a5_val}"

    for row in summary.iter_rows(min_row=2, max_row=summary.max_row, min_col=1, max_col=11):
        for cell in row:
            try:
                if isinstance(cell.value, float) and cell.value.is_integer():
                    cell.value = int(cell.value)
            except:
                pass
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, 12):
        summary.column_dimensions[get_column_letter(col)].width = 15

    wb.save(output_path)

def extract_final_titres_openpyxl(output_path):
    wb = load_workbook(output_path)

    if "Summary" in wb.sheetnames:
        summary = wb["Summary"]
        wb.remove(summary)
    summary = wb.create_sheet("Summary")

    headers = [
        "Plate", "Pseudotype", "Sample ID", 
        "NT 90% Rep 1", "NT 90% Rep 2", "NT 90% Rep 3", "NT 90%",
        "NT 50% Rep 1", "NT 50% Rep 2", "NT 50% Rep 3", "NT 50%"
    ]
    summary.append(headers)

    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("Plate"):
            sheet = wb[sheet_name]

            nt90_cols = [2, 5, 8, 11]
            nt90_row = 14
            nt50_cols = [2, 5, 8, 11]
            nt50_row = 16

            info_cells = []
            for r in range(26, 30):
                row_vals = []
                for c in range(6, 10):
                    val = sheet.cell(row=r, column=c).value
                    row_vals.append(val)
                info_cells.append(row_vals)

            for i, info_row in enumerate(info_cells):
                if len(info_row) < 4:
                    continue

                pseudotype = info_row[0]
                sample_id = info_row[1]
                nt90_total = info_row[2]
                nt50_total = info_row[3]

                nt90_reps = []
                nt50_reps = []

                if i < len(nt90_cols):
                    base_col = nt90_cols[i]
                    for offset in range(3):
                        val = sheet.cell(row=nt90_row, column=base_col + offset).value
                        nt90_reps.append(val if isinstance(val, (int, float)) else "")

                if i < len(nt50_cols):
                    base_col = nt50_cols[i]
                    for offset in range(3):
                        val = sheet.cell(row=nt50_row, column=base_col + offset).value
                        nt50_reps.append(val if isinstance(val, (int, float)) else "")

                summary.append([
                    sheet_name,
                    pseudotype,
                    sample_id,
                    *nt90_reps,
                    nt90_total,
                    *nt50_reps,
                    nt50_total
                ])

    wb.save(output_path)
    add_default_to_final_titres(output_path)

# === Flask routes ===

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/help")
def help_page():
    return render_template("help.html")

@app.route("/settings", methods=["GET", "POST"])
def settings():
    if request.method == "POST":
        timestamp = request.form.get("timestamp_in_title") == "on"
        new_settings = {
            "timestamp_in_title": timestamp,
            "Q1_color": request.form.get("Q1_color", "#ff0000"),
            "Q2_color": request.form.get("Q2_color", "#0000ff"),
            "Q3_color": request.form.get("Q3_color", "#00ff00"),
            "Q4_color": request.form.get("Q4_color", "#800080"),
        }

        file = request.files.get("template_file")
        if file and file.filename.endswith(".xlsx"):
            filename = f"template_{uuid.uuid4().hex}.xlsx"
            filepath = os.path.join("Templates", filename)
            os.makedirs("Templates", exist_ok=True)
            file.save(filepath)
            save_template_path(filepath)
            flash("New template saved and path updated.", "success")

        save_settings(new_settings)
        flash("Settings saved.", "success")
        return redirect(url_for("settings"))

    return render_template("settings.html", settings=load_settings())

@app.route('/process', methods=['POST'])
def process():
    if 'csv_file' not in request.files:
        flash('No file part', 'danger')
        return redirect(url_for('index'))

    csv_file = request.files['csv_file']
    if csv_file.filename == '':
        flash('No selected file', 'danger')
        return redirect(url_for('index'))

    assay_title = request.form.get('assay_title', '').strip()
    num_pseudotypes = int(request.form.get('num_pseudotypes', 1))
    pseudotype_text = request.form.get('pseudotype_text', '').strip()
    sample_id_text = request.form.get('sample_id_text', '').strip()

    if not pseudotype_text:
        flash('Pseudotype(s) must be provided', 'danger')
        return redirect(url_for('index'))
    if not sample_id_text:
        flash('Sample ID(s) must be provided', 'danger')
        return redirect(url_for('index'))

    pseudotypes = [p.strip() for p in pseudotype_text.split(',')]
    sample_ids = [s.strip() for s in sample_id_text.split(',')]

    if len(pseudotypes) < num_pseudotypes:
        flash(f'Expected at least {num_pseudotypes} pseudotypes but got {len(pseudotypes)}', 'danger')
        return redirect(url_for('index'))

    # Save uploaded CSV file temporarily
    temp_dir = tempfile.mkdtemp()
    csv_path = os.path.join(temp_dir, 'input.csv')
    csv_file.save(csv_path)

    # Load template path from config or default path
    try:
        template_path = load_template_path()
    except Exception:
        template_path = os.path.join('static', 'NTA_Template.xlsx')

    # Parse CSV blocks
    blocks = load_csv_blocks(csv_path)

    wb_master = Workbook()
    default_sheet = wb_master.active
    wb_master.remove(default_sheet)

    for i, block in enumerate(blocks, start=1):
        temp_wb = load_workbook(template_path)
        source_ws = temp_wb["Sheet1"]  # change as needed

        # Paste data into source_ws starting B5 (row=5, col=2)
        for r, row_data in enumerate(block, start=5):
            for c, val in enumerate(row_data, start=2):
                source_ws.cell(row=r, column=c, value=val)

        # Fill in assay title, pseudotypes, and sample IDs on this sheet
        fill_plate(source_ws, assay_title, pseudotypes, sample_ids, num_pseudotypes)

        # Create a new worksheet in master workbook and copy all content
        target_ws = wb_master.create_sheet(title=f"Plate {i}")
        copy_worksheet(source_ws, target_ws)

    # Save the combined workbook to temporary path
    output_path = os.path.join(temp_dir, "temp_output.xlsx")
    wb_master.save(output_path)

    # Run your summary extraction function on the output file
    extract_final_titres_openpyxl(output_path)

    # Move final file to upload folder with timestamped filename
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"Processed_{timestamp}.xlsx"
    final_path = os.path.join(app.config["UPLOAD_FOLDER"], output_filename)
    shutil.move(output_path, final_path)

    shutil.rmtree(temp_dir)

    settings = load_settings()

    return render_template('results.html', excel_file=output_filename, plot_file=None, settings=settings)



@app.route("/generate_graphs", methods=["POST"])
def generate_graphs():
    excel_file = request.form.get("excel_file")
    if not excel_file:
        flash("No Excel file found for graph generation.", "danger")
        return redirect(url_for("index"))

    input_path = os.path.join(app.config["UPLOAD_FOLDER"], excel_file)
    output_plot = f"graph_{uuid.uuid4().hex}.png"
    output_path = os.path.join(app.config["UPLOAD_FOLDER"], output_plot)

    include_timestamp = "include_timestamp" in request.form

    q1 = request.form.get("Q1_color", "#ff0000")
    q2 = request.form.get("Q2_color", "#0000ff")
    q3 = request.form.get("Q3_color", "#00ff00")
    q4 = request.form.get("Q4_color", "#800080")

    r_script = os.path.join(os.getcwd(), "process_data.R")

    try:
        subprocess.run(
            [
                "Rscript",
                r_script,
                input_path,
                output_path,
                str(include_timestamp).upper(),
                q1,
                q2,
                q3,
                q4,
            ],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        flash("Graph generated successfully.", "success")

        return render_template(
            "results.html",
            excel_file=excel_file,
            plot_file=output_plot,
            settings=load_settings()
        )

    except subprocess.CalledProcessError as e:
        flash(f"R script failed: {e.stderr}", "danger")
        return redirect(url_for("index"))


@app.route("/download/<filename>")
def download(filename):
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename, as_attachment=True)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
