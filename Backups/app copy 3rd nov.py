from flask import Flask, render_template, request, redirect, url_for, send_file, flash, jsonify, send_from_directory, session
import os
import uuid
import subprocess
from datetime import datetime
from io import BytesIO
import tempfile


from nta_utils import (
    process_csv_to_template,
    extract_final_titres_openpyxl as extract_final_titres_xlwings,
    save_template_path,
    load_template_path,
    load_settings,
    save_settings,
)

in_memory_files = {}  # Key: UUID, Value: BytesIO

app = Flask(__name__)
app.secret_key = "your-secret-key"



@app.route("/")
def index():
    settings = load_settings()
    settings["template_path"] = load_template_path()
    return render_template("index.html", settings=settings)


@app.route("/help")
def help_page():
    return render_template("help.html")

@app.route("/settings", methods=["GET", "POST"])
def settings():
    current_settings = load_settings()

    # Example list of your default templates (absolute or relative paths)
    default_templates = {
        "NTA Template (dil 50-36450)": "excel_templates/NTA_Template.xlsx",
        "Measles Template (dil 32-65536)": "excel_templates/Measles_NTA_Template.xlsx",
        "Backup NTA Template": "excel_templates/Backup_NTA_Template.xlsx",
    }

        # Load the current template path from config.json
    try:
        current_template_path = load_template_path()
    except FileNotFoundError:
        current_template_path = None

    if request.method == "POST":
        # Check if user selected a default template
        selected_template_key = request.form.get("default_template_select")
        if selected_template_key in default_templates:
            selected_template_path = default_templates[selected_template_key]
            save_template_path(selected_template_path)
            flash(f"Template set to {selected_template_key}.", "success")
        else:
            # Otherwise check if user uploaded a new template file
            file = request.files.get("template_file")
            if file and file.filename.endswith(".xlsx"):
                filename = f"template_{uuid.uuid4().hex}.xlsx"
                filepath = os.path.join("excel_templates", filename)
                os.makedirs("excel_templates", exist_ok=True)
                file.save(filepath)
                save_template_path(filepath)
                flash("New template uploaded and path updated.", "success")

        # Save other settings normally here
        timestamp_flag = request.form.get("timestamp_in_filename") == "on"
        new_settings = current_settings.copy()
        new_settings["timestamp_in_filename"] = timestamp_flag
        save_settings(new_settings)
        flash("Settings saved.", "success")
        return redirect(url_for("settings"))
    
    current_settings["template_path"] = current_template_path


    return render_template("settings.html", settings=current_settings, default_templates=default_templates)




@app.route("/process", methods=["POST"])
def process():
    file = request.files["csv_file"]
    if not file:
        flash("No CSV file uploaded.", "danger")
        return redirect(url_for("index"))
    
    assay_title = request.form.get("assay_title", "")
    pseudotypes = request.form.get("pseudotype_text", "").strip()
    sample_ids = request.form.get("sample_id_text", "")

        # Validate pseudotypes input
    if not pseudotypes:
        flash("Please enter at least one pseudotype name.", "danger")
        return redirect(url_for("index"))

    try:
        num_pseudotypes = int(request.form.get("num_pseudotypes", "1"))
        if num_pseudotypes not in [1, 2, 3, 4]:
            raise ValueError()
    except ValueError:
        flash("Invalid pseudotype count. Must be 1–4.", "danger")
        return redirect(url_for("index"))

    settings = load_settings()
    safe_title = assay_title.strip().replace(" ", "_")
    timestamp = datetime.now().strftime("%Y-%m-%d")
    filename = f"{safe_title}_{timestamp}.xlsx" if settings.get("timestamp_in_filename", True) else f"{safe_title}.xlsx"

    csv_bytes = BytesIO(file.read())
    csv_bytes.seek(0)

    try:
        template_path = load_template_path()
    except Exception as e:
        flash(str(e), "danger")
        return redirect(url_for("index"))

    # Create Excel in memory (plate sheets)
    output_bytes = BytesIO()
    process_csv_to_template(
        csv_path=csv_bytes,
        template_path=template_path,
        output_path=output_bytes,
        num_pseudotypes=num_pseudotypes,
        pseudotype_texts=pseudotypes,
        assay_title_text=assay_title,
        sample_id_text=sample_ids,
    )

    # Generate Summary sheet
    extract_final_titres_xlwings(output_bytes)

    # Fill blank NT values with < / > as appropriate
    from nta_utils import add_default_to_final_titres
    add_default_to_final_titres(output_bytes)

    # Save temp file for R plotting
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_excel:
        tmp_excel.write(output_bytes.getvalue())
        excel_path = tmp_excel.name

    # Run R script to generate plot png
    r_script = os.path.join(os.getcwd(), "process_data.R")
    presets = settings.get("presets", {})
    active_preset_name = settings.get("selected_preset", None)
    default_colours = {"Q1": "#ff7e79", "Q2": "#ffd479", "Q3": "#009193", "Q4": "#d783ff"}
    colours = presets.get(active_preset_name, default_colours)

    q1_colour = colours.get("Q1", default_colours["Q1"])
    q2_colour = colours.get("Q2", default_colours["Q2"])
    q3_colour = colours.get("Q3", default_colours["Q3"])
    q4_colour = colours.get("Q4", default_colours["Q4"])

    quadrants = settings.get("quadrants", {"Q1": True, "Q2": True, "Q3": True, "Q4": True})
    q1_flag = str(quadrants.get("Q1", True)).lower()
    q2_flag = str(quadrants.get("Q2", True)).lower()
    q3_flag = str(quadrants.get("Q3", True)).lower()
    q4_flag = str(quadrants.get("Q4", True)).lower()

    plot_title = os.path.splitext(filename)[0]
    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_png:
        output_plot_path = tmp_png.name

    subprocess.run(
        [
            "Rscript",
            r_script,
            excel_path,
            output_plot_path,
            str(settings.get("timestamp_in_filename", True)).lower(),
            q1_colour,
            q2_colour,
            q3_colour,
            q4_colour,
            plot_title,
            q1_flag, q2_flag, q3_flag, q4_flag
        ],
        check=True
    )

    # Embed graphs into Excel
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage

    wb = load_workbook(excel_path)

    # --- Summary graph on its own sheet ---
    ws = wb.create_sheet("Summary Plots")
    img = XLImage(output_plot_path)
    ws.add_image(img, "A1")

    # --- Per-plate graphs ---
    plate_dir = os.path.dirname(output_plot_path)
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Plate"):
            continue

        plate_png = os.path.join(plate_dir, f"{sheet_name}.png")
        if os.path.exists(plate_png):
            ws_plate = wb[sheet_name]

            # ensure Excel-compatible PNG
            pil_img = PILImage.open(plate_png)
            temp_png = os.path.join(plate_dir, f"temp_{sheet_name}.png")
            pil_img.save(temp_png, "PNG")

            img_plate = XLImage(temp_png)
            img_plate.anchor = "B33"  # place to the right of data
            ws_plate.add_image(img_plate)

    wb.save(excel_path)

    # Load back into BytesIO for download
    with open(excel_path, "rb") as f:
        final_bytes = BytesIO(f.read())

    # Cleanup temp files
    os.remove(excel_path)
    os.remove(output_plot_path)

    file_id = uuid.uuid4().hex
    in_memory_files[file_id] = {"data": final_bytes.getvalue(), "name": filename}
    session["file_id"] = file_id

    return render_template(
        "results.html",
        excel_file_id=file_id,
        filename=filename,
        settings=settings,
        plot_file=None
    )


@app.route("/download_memory/<file_id>")
def download_memory(file_id):
    file_info = in_memory_files.get(file_id)
    if not file_info:
        flash("File not found in memory.", "danger")
        return redirect(url_for("index"))

    # Create fresh BytesIO for each request
    file_stream = BytesIO(file_info["data"])
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name=file_info["name"],
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/generate_graphs", methods=["POST"])
def generate_graphs():
    file_id = request.form.get("file_id")
    if not file_id or file_id not in in_memory_files:
        flash("No Excel file found for graph generation.", "danger")
        return redirect(url_for("index"))

    file_info = in_memory_files[file_id]
    file_bytes = file_info["data"]
    filename = file_info["name"]

    r_script = os.path.join(os.getcwd(), "process_data.R")

    settings = load_settings()
    include_timestamp = settings.get("timestamp_in_filename", True)

    presets = settings.get("presets", {})
    active_preset_name = settings.get("selected_preset", None)
    default_colours = {
        "Q1": "#ff7e79", "Q2": "#ffd479", "Q3": "#009193", "Q4": "#d783ff"
    }
    colours = presets.get(active_preset_name, default_colours)

    q1_colour = colours.get("Q1", default_colours["Q1"])
    q2_colour = colours.get("Q2", default_colours["Q2"])
    q3_colour = colours.get("Q3", default_colours["Q3"])
    q4_colour = colours.get("Q4", default_colours["Q4"])
    quadrants = settings.get("quadrants", {"Q1": True, "Q2": True, "Q3": True, "Q4": True})
    q1_flag = quadrants.get("Q1", True)
    q2_flag = quadrants.get("Q2", True)
    q3_flag = quadrants.get("Q3", True)
    q4_flag = quadrants.get("Q4", True)

    try:
        # Create fresh BytesIO from bytes
        file_stream = BytesIO(file_bytes)
        file_stream.seek(0)

        # Save Excel file to temporary path for Rscript
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_input:
            tmp_input.write(file_stream.read())
            input_path = tmp_input.name

        # Temporary PNG output path
        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_output:
            output_plot_path = tmp_output.name

        # Run R script
        plot_title = os.path.splitext(filename)[0]
        subprocess.run(
            [
                "Rscript",
                r_script,
                input_path,
                output_plot_path,
                str(include_timestamp).lower(),
                q1_colour,
                q2_colour,
                q3_colour,
                q4_colour,
                plot_title,
                str(q1_flag).lower(),
                str(q2_flag).lower(),
                str(q3_flag).lower(),
                str(q4_flag).lower()
            ],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )

        # Load PNG image into memory
        with open(output_plot_path, "rb") as f: 
            image_bytes = BytesIO(f.read())

        image_bytes.seek(0)

        # Cleanup temp files
        os.remove(input_path)
        os.remove(output_plot_path)

        # Return PNG as download
        png_filename = os.path.splitext(filename)[0] + ".png"
        return send_file(
            image_bytes,
            mimetype="image/png",
            as_attachment=True,
            download_name=png_filename
        )

    except subprocess.CalledProcessError as e:
        flash(f"R script failed: {e.stderr}", "danger")
        return redirect(url_for("index"))

    except Exception as e:
        flash(f"Unexpected error: {str(e)}", "danger")
        return redirect(url_for("index"))


@app.route("/save_quadrants", methods=["POST"])
def save_quadrants():
    quadrants = request.get_json()
    settings = load_settings()
    settings["quadrants"] = quadrants
    save_settings(settings)
    return "Quadrant settings saved", 200


@app.route("/get_settings")
def get_settings():
    settings = load_settings()
    return jsonify(settings)


@app.route("/save_preset", methods=["POST"])
def save_preset():
    data = request.get_json()
    settings = load_settings()
    settings["presets"][data["name"]] = data["colours"]
    save_settings(settings)
    return jsonify({"status": "success", "message": "Preset saved."}), 200
    

@app.route("/delete_preset", methods=["POST"])
def delete_preset():
    data = request.get_json()
    settings = load_settings()
    settings["presets"].pop(data["name"], None)
    save_settings(settings)
    return jsonify({"status": "success", "message": "Preset saved."}), 200


@app.route("/set_active_preset", methods=["POST"])
def set_active_preset():
    data = request.get_json()
    name = data.get("name")
    if not name:
        return "No preset name provided", 400

    settings = load_settings()
    if "presets" not in settings or name not in settings["presets"]:
        return "Preset not found", 404

    settings["selected_preset"] = name
    save_settings(settings)

    return "Preset updated", 200

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)