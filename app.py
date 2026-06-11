from flask import Flask, render_template, request, redirect, url_for, send_file, flash, jsonify, send_from_directory, session
import os
import uuid
import json
import logging
import subprocess
import threading
import time
from datetime import datetime
from io import BytesIO
import re
import tempfile
from itertools import combinations
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from PIL import Image as PILImage


from nta_utils import (
    process_csv_to_template,
    extract_final_titres_openpyxl as extract_final_titres_xlwings,
    save_template_path,
    load_template_path,
    load_settings,
    save_settings,
    generate_sigmoid_csv,
    flag_triplicate_errors,
    count_errors_from_workbook,
    validate_csv_mode,
    DEFAULT_SETTINGS,
)

in_memory_files = {}  # Key: UUID, Value: BytesIO

app = Flask(__name__)
app.secret_key = "your-secret-key"
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB upload limit

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("ntaweb")



@app.route("/")
def index():
    settings = load_settings()
    settings["template_path"] = load_template_path()
    return render_template("index.html", settings=settings)


@app.route("/help")
def help_page():
    return "", 204

@app.route("/save_timestamp_setting", methods=["POST"])
def save_timestamp_setting():
    data = request.get_json()
    settings = load_settings()
    settings["timestamp_in_filename"] = bool(data.get("enabled", False))
    save_settings(settings)
    return jsonify({"status": "ok"})

@app.route("/settings", methods=["GET", "POST"])
def settings():
    current_settings = load_settings()

    default_templates = {
        "NTA Template (dil 50-36450)": "excel_templates/NTA_Template.xlsx",
        "Measles Template (dil 32-65536)": "excel_templates/Measles_NTA_Template.xlsx",
        "Backup NTA Template": "excel_templates/Backup_NTA_Template.xlsx",
    }
    # Append user-created templates (only if the file still exists)
    for name, path in current_settings.get("custom_templates", {}).items():
        if os.path.exists(path):
            default_templates[name] = path

    try:
        current_template_path = load_template_path()
    except FileNotFoundError:
        current_template_path = None

    if request.method == "POST":
        selected_template_key = request.form.get("default_template_select")
        if selected_template_key in default_templates:
            selected_template_path = default_templates[selected_template_key]
            save_template_path(selected_template_path)
        else:
            file = request.files.get("template_file")
            if file and file.filename.endswith(".xlsx"):
                filename = f"template_{uuid.uuid4().hex}.xlsx"
                filepath = os.path.join("excel_templates", filename)
                os.makedirs("excel_templates", exist_ok=True)
                file.save(filepath)
                save_template_path(filepath)

        timestamp_flag = request.form.get("timestamp_in_filename") == "on"
        error_flagging_flag = request.form.get("error_flagging") == "on"
        new_settings = current_settings.copy()
        new_settings["timestamp_in_filename"] = timestamp_flag
        new_settings["error_flagging"] = error_flagging_flag
        new_settings["default_data_mode"] = request.form.get("default_data_mode", "standard")
        try:
            new_settings["default_num_pseudotypes"] = int(request.form.get("default_num_pseudotypes", 1))
        except ValueError:
            new_settings["default_num_pseudotypes"] = 1
        try:
            new_settings["outlier_threshold_log2"] = float(request.form.get("outlier_threshold_log2", 1.0))
        except ValueError:
            new_settings["outlier_threshold_log2"] = 1.0
        try:
            new_settings["sigmoid_r2_threshold"] = float(request.form.get("sigmoid_r2_threshold", 0.5))
        except ValueError:
            new_settings["sigmoid_r2_threshold"] = 0.5
        new_settings["lod_censor_include"] = request.form.get("lod_censor_include") == "on"
        try:
            new_settings["comparison_disagreement_threshold"] = float(request.form.get("comparison_disagreement_threshold", 1.0))
        except ValueError:
            new_settings["comparison_disagreement_threshold"] = 1.0
        save_settings(new_settings)
        flash("Settings saved.", "success")
        return redirect(url_for("settings"))

    current_settings["template_path"] = current_template_path

    return render_template("settings.html", settings=current_settings, default_templates=default_templates)


@app.route("/reset_settings", methods=["POST"])
def reset_settings():
    """Reset threshold/toggle settings to defaults, preserving presets and custom templates."""
    current = load_settings()
    reset_keys = [
        "timestamp_in_filename", "error_flagging", "default_data_mode",
        "default_num_pseudotypes", "outlier_threshold_log2",
        "sigmoid_r2_threshold", "lod_censor_include", "comparison_disagreement_threshold",
    ]
    for key in reset_keys:
        current[key] = DEFAULT_SETTINGS[key]
    save_settings(current)
    flash("Settings reset to defaults.", "success")
    return redirect(url_for("settings"))


@app.route("/save_template_selection", methods=["POST"])
def save_template_selection():
    """Lightweight endpoint for the template selector + upload only."""
    current_settings = load_settings()
    default_templates = {
        "NTA Template (dil 50-36450)": "excel_templates/NTA_Template.xlsx",
        "Measles Template (dil 32-65536)": "excel_templates/Measles_NTA_Template.xlsx",
        "Backup NTA Template": "excel_templates/Backup_NTA_Template.xlsx",
    }
    for name, path in current_settings.get("custom_templates", {}).items():
        if os.path.exists(path):
            default_templates[name] = path

    selected_key = request.form.get("default_template_select")
    if selected_key in default_templates:
        save_template_path(default_templates[selected_key])
    else:
        file = request.files.get("template_file")
        if file and file.filename.endswith(".xlsx"):
            filename = f"template_{uuid.uuid4().hex}.xlsx"
            filepath = os.path.join("excel_templates", filename)
            os.makedirs("excel_templates", exist_ok=True)
            file.save(filepath)
            save_template_path(filepath)

    flash("Template saved.", "success")
    return redirect(url_for("settings"))


def _run_r_in_background(file_id, excel_path, output_plot_path, r_cmd):
    """Run process_data.R in a background thread, then embed plots into the stored Excel."""
    try:
        subprocess.run(r_cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        logger.info("R SCRIPT (background) complete for %s", file_id)

        with open(output_plot_path, "rb") as f:
            summary_plot_bytes = f.read()

        # Embed plots into the stored Excel bytes
        file_info = in_memory_files.get(file_id)
        if file_info:
            wb = load_workbook(BytesIO(file_info["data"]))
            ws_summary = wb.create_sheet("Summary Plots")
            ws_summary.add_image(XLImage(output_plot_path), "A1")

            plate_dir = os.path.dirname(output_plot_path)
            _plates_embedded = 0
            for sheet_name in wb.sheetnames:
                if not sheet_name.startswith("Plate"):
                    continue
                plate_png = os.path.join(plate_dir, f"{sheet_name}.png")
                if os.path.exists(plate_png):
                    ws_plate = wb[sheet_name]
                    temp_png = os.path.join(plate_dir, f"temp_{sheet_name}.png")
                    PILImage.open(plate_png).save(temp_png, "PNG")
                    img_plate = XLImage(temp_png)
                    img_plate.anchor = "B33"
                    ws_plate.add_image(img_plate)
                    _plates_embedded += 1
            logger.info("IMAGES   (background) %d plate PNG(s) embedded", _plates_embedded)

            out = BytesIO()
            wb.save(out)
            file_info["data"] = out.getvalue()
            file_info["summary_plot"] = summary_plot_bytes
            file_info["plots_ready"] = True
            logger.info("PLOTS    stored for %s", file_id)
    except Exception:
        logger.exception("R SCRIPT (background) failed for %s", file_id)
        if file_id in in_memory_files:
            in_memory_files[file_id]["plots_ready"] = True  # stop polling, fall back to on-demand
    finally:
        try:
            os.remove(excel_path)
        except OSError:
            pass
        try:
            os.remove(output_plot_path)
        except OSError:
            pass
        plate_dir = os.path.dirname(output_plot_path)
        for f in os.listdir(plate_dir):
            if f.endswith(".png"):
                try:
                    os.remove(os.path.join(plate_dir, f))
                except OSError:
                    pass


@app.route("/process", methods=["POST"])
def process():
    file = request.files["csv_file"]
    if not file:
        flash("No CSV file uploaded.", "danger")
        return redirect(url_for("index"))

    assay_title = request.form.get("assay_title", "")
    pseudotypes = request.form.get("pseudotype_text", "").strip()
    sample_ids = request.form.get("sample_id_text", "")
    data_mode = request.form.get("data_mode", "standard")

    if data_mode not in ("data_only", "standard"):
        data_mode = "standard"

    if not pseudotypes:
        flash("Please enter at least one pseudotype name.", "danger")
        return redirect(url_for("index"))

    raw_np = request.form.get("num_pseudotypes", "1")
    if raw_np == "2alt":
        num_pseudotypes = "2alt"
    else:
        try:
            num_pseudotypes = int(raw_np)
            if num_pseudotypes not in [1, 2, 3, 4]:
                raise ValueError()
        except ValueError:
            flash("Invalid pseudotype count. Must be 1–4.", "danger")
            return redirect(url_for("index"))

    # Per-plate config (optional — sent as JSON when custom per-plate mode is active)
    plate_configs = None
    raw_pc = request.form.get("plate_configs", "").strip()
    if raw_pc:
        try:
            plate_configs = json.loads(raw_pc)
            if not isinstance(plate_configs, list):
                plate_configs = None
        except (ValueError, TypeError):
            plate_configs = None

    settings = load_settings()
    safe_title = assay_title.strip().replace(" ", "_")
    timestamp = datetime.now().strftime("%Y-%m-%d")
    filename = f"{safe_title}_{timestamp}.xlsx" if settings.get("timestamp_in_filename", True) else f"{safe_title}.xlsx"

    csv_bytes = BytesIO(file.read())
    csv_bytes.seek(0)

    try:
        template_path = load_template_path()
    except Exception as e:
        flash(str(e), "danger")
        return redirect(url_for("index"))

    csv_size_kb = round(csv_bytes.getbuffer().nbytes / 1024, 1)
    logger.info("PROCESS  \u2190 %r  mode=%s  pseudotypes=%s", assay_title, data_mode, pseudotypes.replace("\n", ","))
    logger.info("CSV      read %.1f KB", csv_size_kb)

    _proc_start = time.time()
    _t = time.time()
    output_bytes = BytesIO()
    process_csv_to_template(
        csv_path=csv_bytes,
        template_path=template_path,
        output_path=output_bytes,
        num_pseudotypes=num_pseudotypes,
        pseudotype_texts=pseudotypes,
        assay_title_text=assay_title,
        sample_id_text=sample_ids,
        data_mode=data_mode,
        plate_configs=plate_configs,
    )
    logger.info("EXCEL    workbook built in %.1fs", time.time() - _t)

    _t = time.time()
    extract_final_titres_xlwings(output_bytes)

    from nta_utils import add_default_to_final_titres
    add_default_to_final_titres(output_bytes)
    logger.info("EXTRACT  final titres written in %.1fs", time.time() - _t)

    # ── Error flagging (if enabled in settings) ──
    if settings.get("error_flagging", False):
        _t = time.time()
        flag_triplicate_errors(output_bytes, threshold_log2=settings.get("outlier_threshold_log2", 1.0))
        from nta_utils import count_errors_from_workbook
        _err_count, _ = count_errors_from_workbook(output_bytes.getvalue())
        logger.info("FLAGS    error flagging complete in %.1fs (%d flagged)", time.time() - _t, _err_count)
    else:
        logger.info("FLAGS    disabled")

    # Store Excel immediately (no plots yet) so we can respond without waiting for R
    file_id = uuid.uuid4().hex
    in_memory_files[file_id] = {
        "data": output_bytes.getvalue(),
        "name": filename,
        "summary_plot": None,
        "plots_ready": False,
    }
    session["file_id"] = file_id

    # Build R command args \u2014 R runs in background so temp files must persist until it finishes
    r_script = os.path.join(os.getcwd(), "process_data.R")
    presets = settings.get("presets", {})
    active_preset_name = settings.get("selected_preset", None)
    default_colours = {"Q1": "#ff7e79", "Q2": "#ffd479", "Q3": "#009193", "Q4": "#d783ff"}
    colours = presets.get(active_preset_name, default_colours)

    q1_colour = colours.get("Q1", default_colours["Q1"])
    q2_colour = colours.get("Q2", default_colours["Q2"])
    q3_colour = colours.get("Q3", default_colours["Q3"])
    q4_colour = colours.get("Q4", default_colours["Q4"])

    quadrants = settings.get("quadrants", {"Q1": True, "Q2": True, "Q3": True, "Q4": True})
    q1_flag = str(quadrants.get("Q1", True)).lower()
    q2_flag = str(quadrants.get("Q2", True)).lower()
    q3_flag = str(quadrants.get("Q3", True)).lower()
    q4_flag = str(quadrants.get("Q4", True)).lower()

    plot_title = os.path.splitext(filename)[0]

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_excel:
        tmp_excel.write(output_bytes.getvalue())
        excel_path = tmp_excel.name

    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_png:
        output_plot_path = tmp_png.name

    r_cmd = [
        "Rscript", r_script,
        excel_path, output_plot_path,
        str(settings.get("timestamp_in_filename", True)).lower(),
        q1_colour, q2_colour, q3_colour, q4_colour,
        plot_title,
        q1_flag, q2_flag, q3_flag, q4_flag,
    ]

    logger.info("R SCRIPT launching in background for %s", file_id)
    threading.Thread(
        target=_run_r_in_background,
        args=(file_id, excel_path, output_plot_path, r_cmd),
        daemon=True,
    ).start()

    _proc_elapsed = round(time.time() - _proc_start, 1)
    logger.info("DONE     \u2713 %r ready (plots pending) \u00b7 %.1fs", filename, _proc_elapsed)
    return render_template(
        "analysis_hub.html",
        excel_file_id=file_id,
        filename=filename,
        processing_time=_proc_elapsed,
        fitting_id=None,
        settings=load_settings(),
    )


# ════════════════════════════════════════════════════════════════
# Data Analysis
# ════════════════════════════════════════════════════════════════

@app.route("/hub/<file_id>")
def analysis_hub(file_id):
    """Data Analysis — the central page with three analysis options."""
    if file_id not in in_memory_files:
        flash("Results not found. Please process your data again.", "danger")
        return redirect(url_for("index"))

    file_info = in_memory_files[file_id]
    return render_template(
        "analysis_hub.html",
        excel_file_id=file_id,
        filename=file_info.get("name", "results.xlsx"),
        fitting_id=file_info.get("fitting_id"),
        comparison_id=file_info.get("comparison_id"),
        settings=load_settings(),
    )


@app.route("/linear/<file_id>")
def linear_results(file_id):
    """Dedicated Linear Interpolation results page."""
    if file_id not in in_memory_files:
        flash("Results not found. Please process your data again.", "danger")
        return redirect(url_for("index"))

    file_info = in_memory_files[file_id]
    return render_template(
        "linear_results.html",
        excel_file_id=file_id,
        filename=file_info.get("name", "results.xlsx"),
        settings=load_settings(),
    )


@app.route("/linear_summary/<file_id>")
def linear_summary_data(file_id):
    """JSON API returning plate/pseudotype/sample/titre counts for Data Summary card."""
    if file_id not in in_memory_files:
        return jsonify({"status": "error", "message": "File not found"})

    try:
        file_info = in_memory_files[file_id]

        # Return cached result to avoid re-parsing the workbook on every page load
        if "_summary_cache" in file_info:
            return jsonify(file_info["_summary_cache"])

        file_bytes = file_info["data"]
        wb = load_workbook(BytesIO(file_bytes), data_only=True)

        plate_sheets = [s for s in wb.sheetnames if s.startswith("Plate")]

        # Only count plates that contain actual numeric well data (B5:M12).
        # This excludes any extra/blank plates the plate reader appended.
        def _plate_has_data(ws):
            for row in range(5, 13):
                for col in ['B','C','D','E','F','G','H','I','J','K','L','M']:
                    try:
                        float(ws[f'{col}{row}'].value)
                        return True
                    except (ValueError, TypeError):
                        pass
            return False

        num_plates = sum(1 for s in plate_sheets if _plate_has_data(wb[s]))

        pseudotypes = set()
        num_quadrants = 0
        num_labelled = 0
        has_any_label = False
        all_labelled = True

        for sheet_name in plate_sheets:
            ws = wb[sheet_name]
            for pt_cell, sid_cell in [('B3','B4'), ('E3','E4'), ('H3','H4'), ('K3','K4')]:
                pt_val = ws[pt_cell].value
                sid_val = ws[sid_cell].value
                if pt_val and str(pt_val).strip():
                    pseudotypes.add(str(pt_val).strip())
                    num_quadrants += 1
                    if sid_val and str(sid_val).strip():
                        has_any_label = True
                        num_labelled += 1
                    else:
                        all_labelled = False

        # Determine labelling status
        if num_quadrants == 0:
            label_status = "none"
        elif all_labelled:
            label_status = "labelled"
        elif has_any_label:
            label_status = "partial"
        else:
            label_status = "unlabelled"

        # Count flagged errors (if Errors sheet exists)
        error_count, has_errors_sheet = count_errors_from_workbook(file_bytes)

        result = {
            "status": "success",
            "num_plates": num_plates,
            "num_pseudotypes": len(pseudotypes),
            "num_samples": num_quadrants,
            "num_labelled": num_labelled,
            "label_status": label_status,
            "error_count": error_count,
            "error_flagging_enabled": has_errors_sheet,
        }
        file_info["_summary_cache"] = result
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


def _compute_boxplot_data(file_bytes, threshold_pct):
    """
    Compute NT titres for box plot using the same formula as the Excel template.

    For each replicate column (3 per quadrant) on each Plate sheet:
      NSC     = luminescence at row 12 (no-serum control, index 7 in 0-based)
      target  = NSC × (1 - threshold_pct/100)  e.g. NT50 → NSC × 0.5
      P16     = last 0-based index (0–7) where lum ≤ target
                (replicates Excel MATCH(target, col5:col12, 1) on ascending data)

      Boundary conditions:
        P16 is None  → lum never drops to target → NT ≤ A5  (lower boundary)
        P16 = 6      → crossing involves NSC row  → NT ≥ A11 (upper boundary)
        P16 in 0–5   → valid interpolation between two tested dilution points

      NT (valid) = (target - lum[P16]) / (lum[P16+1] - lum[P16])
                   × (dil[P16+1] - dil[P16]) + dil[P16]

    Each entry stores both:
      "nt"          — average of non-boundary replicates only (used when boundary toggle OFF)
      "nt_boundary" — average substituting A5 for low-boundary and A11 for high-boundary
                      replicates (used when boundary toggle ON)

    Returns:
      {
        pseudotype: [
          { sample, plate, nt, nt_boundary, has_boundary,
            boundary_low, boundary_high, nsc, target,
            rep_nts, rep_boundary_flags },
          ...
        ]
      }
    """
    target_fraction = (100 - threshold_pct) / 100

    wb = load_workbook(BytesIO(file_bytes), data_only=True)

    quad_defs = [
        {"pt_cell": "B3", "sid_cell": "B4", "cols": ["B", "C", "D"]},
        {"pt_cell": "E3", "sid_cell": "E4", "cols": ["E", "F", "G"]},
        {"pt_cell": "H3", "sid_cell": "H4", "cols": ["H", "I", "J"]},
        {"pt_cell": "K3", "sid_cell": "K4", "cols": ["K", "L", "M"]},
    ]

    plate_sheets = sorted(
        [s for s in wb.sheetnames if re.match(r"^Plate\d+$", s)],
        key=lambda s: int(s[5:]),
    )

    grouped = {}

    for sheet_name in plate_sheets:
        ws = wb[sheet_name]

        # Dilutions A5:A12 (8 values)
        # Index 0 = A5 (lowest tested dilution)
        # Index 6 = A11 (highest tested dilution)
        # Index 7 = A12 (NSC slot — not a real dilution point)
        dilutions = []
        for row in range(5, 13):
            val = ws[f"A{row}"].value
            try:
                dilutions.append(float(val))
            except (ValueError, TypeError):
                dilutions.append(None)

        dil_low  = dilutions[0]  # A5  — lower boundary limit
        dil_high = dilutions[6]  # A11 — upper boundary limit

        for quad in quad_defs:
            pt_val = ws[quad["pt_cell"]].value
            if not pt_val or not str(pt_val).strip():
                continue

            pseudotype = str(pt_val).strip()
            sid_val    = ws[quad["sid_cell"]].value
            sample     = str(sid_val).strip() if sid_val and str(sid_val).strip() else "Unlabelled"

            # rep_data: list of (nt_valid_or_None, boundary_flag)
            # boundary_flag: None = valid, "low" = ≤A5, "high" = ≥A11
            rep_data = []

            for col in quad["cols"]:
                lum = []
                for row in range(5, 13):
                    val = ws[f"{col}{row}"].value
                    try:
                        lum.append(float(val))
                    except (ValueError, TypeError):
                        lum.append(None)

                nsc = lum[7]
                if nsc is None or nsc <= 0:
                    rep_data.append((None, None))
                    continue

                target = nsc * target_fraction

                # Excel MATCH(target, lum, 1): binary search (assumes ascending)
                lo, hi = 0, 6  # search lum[0:6] (rows 5-11, not NSC)
                p16 = None
                while lo <= hi:
                    mid = (lo + hi) // 2
                    if lum[mid] is not None and lum[mid] <= target:
                        lo = mid + 1
                    else:
                        hi = mid - 1
                if hi >= 0:
                    p16 = hi

                if p16 is None:
                    # All lum[0:6] > target → NT ≤ lowest dilution
                    rep_data.append((None, "low"))
                else:
                    # p16=6: uses (lum[6], lum[7]=NSC) / (dilutions[6], dilutions[7]=0)
                    y1, y2 = lum[p16], lum[p16 + 1]
                    x1, x2 = dilutions[p16], dilutions[p16 + 1]
                    if None in (y1, y2, x1, x2) or y2 == y1:
                        rep_data.append((None, None))
                    else:
                        nt_val = (target - y1) / (y2 - y1) * (x2 - x1) + x1
                        rep_data.append((nt_val if nt_val > 0 else None, None))

            # Average without boundary substitution
            valid_nts = [nt for nt, b in rep_data if nt is not None and b is None]

            # Average with boundary substitution (A5 for low, A11 for high)
            boundary_nts = []
            for nt, b in rep_data:
                if b is None and nt is not None:
                    boundary_nts.append(nt)
                elif b == "low"  and dil_low  is not None:
                    boundary_nts.append(dil_low)
                elif b == "high" and dil_high is not None:
                    boundary_nts.append(dil_high)

            # Skip entries with no data at all
            if not valid_nts and not boundary_nts:
                continue

            avg_nt          = round(sum(valid_nts)    / len(valid_nts),    1) if valid_nts    else None
            avg_nt_boundary = round(sum(boundary_nts) / len(boundary_nts), 1) if boundary_nts else None

            first_col = quad["cols"][0]
            ref_nsc = None
            try:
                ref_nsc = float(ws[f"{first_col}12"].value)
            except (ValueError, TypeError):
                pass
            ref_target = round(ref_nsc * target_fraction, 2) if ref_nsc else None

            entry = {
                "sample":             sample,
                "plate":              sheet_name,
                "nt":                 avg_nt,           # boundary-excluded average
                "nt_boundary":        avg_nt_boundary,  # boundary-included average
                "has_boundary":       any(b is not None for _, b in rep_data),
                "boundary_low":       dil_low,
                "boundary_high":      dil_high,
                "nsc":                round(ref_nsc, 2) if ref_nsc else None,
                "target":             ref_target,
                "rep_nts":            [round(nt, 1) if nt is not None else None for nt, _ in rep_data],
                "rep_boundary_flags": [b for _, b in rep_data],
            }
            grouped.setdefault(pseudotype, []).append(entry)

    return grouped


@app.route("/boxplot_data/<file_id>")
def boxplot_data(file_id):
    """JSON API: computes NT titres using the same formula as the Excel template
    (linear interpolation per replicate, averaged across triplicates),
    grouped by pseudotype for the box plot.

    Query params:
        threshold: 50 (default) or 90
    """
    if file_id not in in_memory_files:
        return jsonify({"status": "error", "message": "File not found"})

    threshold = request.args.get("threshold", "50")
    if threshold not in ("50", "90"):
        threshold = "50"

    include_boundary = request.args.get("boundary", "false").lower() == "true"

    q_active = {
        "Q1": request.args.get("q1", "true").lower() != "false",
        "Q2": request.args.get("q2", "true").lower() != "false",
        "Q3": request.args.get("q3", "true").lower() != "false",
        "Q4": request.args.get("q4", "true").lower() != "false",
    }

    try:
        file_info = in_memory_files[file_id]
        file_bytes = file_info["data"]

        # ── Cache hit (raw data — boundary filtering applied per-request) ──
        bp_cache = file_info.get("boxplot_cache", {})
        if threshold in bp_cache:
            logger.info("BOXPLOT  NT%s — cache hit", threshold)
            raw_grouped = bp_cache[threshold]["data"]
        else:
            # ── Compute ──────────────────────────────────────────────
            logger.info("BOXPLOT  NT%s — computing …", threshold)
            _t = time.time()
            raw_grouped = _compute_boxplot_data(file_bytes, int(threshold))
            logger.info("BOXPLOT  NT%s — done in %.2fs", threshold, time.time() - _t)
            file_info.setdefault("boxplot_cache", {})[threshold] = {
                "titre_label": f"NT{threshold}",
                "data":        raw_grouped,
            }

        # ── Apply boundary mode: pick which NT value to use ──────────
        # Build a fresh view — never mutate the cached dicts
        nt_key = "nt_boundary" if include_boundary else "nt"
        filtered = {}
        for pt, entries in raw_grouped.items():
            kept = []
            for e in entries:
                nt = e.get(nt_key)
                if nt is None:
                    continue
                view = dict(e)
                view["nt"] = nt          # normalise: frontend always reads "nt"
                view["include_boundary"] = include_boundary
                kept.append(view)
            if kept:
                filtered[pt] = kept

        # ── Filter to active quadrants ───────────────────────────────
        q_pt_cells = {"Q1": "B3", "Q2": "E3", "Q3": "H3", "Q4": "K3"}
        wb_f = load_workbook(BytesIO(file_bytes), data_only=True)
        allowed = set()
        for sheet in [s for s in wb_f.sheetnames if re.match(r"^Plate\d+$", s)]:
            ws_f = wb_f[sheet]
            for q, cell in q_pt_cells.items():
                if q_active[q]:
                    val = ws_f[cell].value
                    if val and str(val).strip():
                        allowed.add(str(val).strip())
        if allowed:
            filtered = {k: v for k, v in filtered.items() if k in allowed}

        return jsonify({
            "status":      "success",
            "titre_label": f"NT{threshold}",
            "data":        filtered,
        })

    except Exception as e:
        logger.exception("BOXPLOT  error")
        return jsonify({"status": "error", "message": str(e)})


@app.route("/curve_fitting_results/<fitting_id>")
def curve_fitting_results(fitting_id):
    """Serve cached sigmoid curve fitting results without reprocessing."""
    if fitting_id not in in_memory_files:
        flash("Curve fitting results not found. Please run curve fitting again.", "danger")
        return redirect(url_for("index"))
    info = in_memory_files[fitting_id]
    excel_file_id = info.get("excel_file_id")
    if not excel_file_id or excel_file_id not in in_memory_files:
        flash("Original Excel results not found.", "danger")
        return redirect(url_for("index"))
    return render_template(
        "curve_fitting_results.html",
        fitting_id=fitting_id,
        excel_file_id=excel_file_id,
        output_files=list(info["data"].keys()),
        ic50_filename=info["ic50_filename"],
        lod_used=info.get("include_lod", load_settings().get("lod_censor_include", False)),
        settings=load_settings(),
        processing_time=None,
    )


@app.route("/compare_titres_page/<file_id>")
def compare_titres_page(file_id):
    """
    Titre comparison — triggered from Data Analysis.
    Expects ?fitting_id=<id> in query string.
    Immediately runs the comparison and shows results.
    """
    fitting_id = request.args.get("fitting_id")
    if not file_id or file_id not in in_memory_files:
        flash("Excel results not found.", "danger")
        return redirect(url_for("index"))
    if not fitting_id or fitting_id not in in_memory_files:
        flash("Curve fitting results not found. Please perform curve fitting first.", "danger")
        return redirect(url_for("analysis_hub", file_id=file_id))

    # Return cached comparison results if already computed for this file
    existing_cmp_id = in_memory_files[file_id].get("comparison_id")
    if existing_cmp_id and existing_cmp_id in in_memory_files:
        cached = in_memory_files[existing_cmp_id]
        return render_template(
            "titre_comparison_results.html",
            comparison_id=existing_cmp_id,
            excel_file_id=file_id,
            stats=cached["stats"],
            mismatches=cached["mismatches"],
            has_plot=cached["has_plot"],
            settings=load_settings(),
            processing_time=None,
        )

    # Delegate to the existing compare logic
    return _run_comparison(file_id, fitting_id)


# ════════════════════════════════════════════════════════════════
# download / utility routes
# ════════════════════════════════════════════════════════════════

@app.route("/download_memory/<file_id>")
def download_memory(file_id):
    file_info = in_memory_files.get(file_id)
    if not file_info:
        flash("File not found in memory.", "danger")
        return redirect(url_for("index"))

    # Block until background R thread finishes embedding plots (max 120s)
    deadline = time.time() + 120
    while not file_info.get("plots_ready") and time.time() < deadline:
        time.sleep(1)

    file_stream = BytesIO(file_info["data"])
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name=file_info["name"],
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/generate_graphs", methods=["POST"])
def generate_graphs():
    file_id = request.form.get("file_id")
    if not file_id or file_id not in in_memory_files:
        flash("No Excel file found for graph generation.", "danger")
        return redirect(url_for("index"))

    file_info = in_memory_files[file_id]
    file_bytes = file_info["data"]
    filename = file_info["name"]

    r_script = os.path.join(os.getcwd(), "process_data.R")

    settings = load_settings()
    include_timestamp = settings.get("timestamp_in_filename", True)
    default_colours = {"Q1": "#ff7e79", "Q2": "#ffd479", "Q3": "#009193", "Q4": "#d783ff"}

    graph_preset    = request.form.get("graph_preset", "").strip()
    graph_quadrants = request.form.get("graph_quadrants", "").strip()

    if graph_preset and graph_preset in settings.get("presets", {}):
        colours = settings["presets"][graph_preset]
    else:
        active_preset_name = settings.get("selected_preset", None)
        colours = settings.get("presets", {}).get(active_preset_name, default_colours)

    q1_colour = colours.get("Q1", default_colours["Q1"])
    q2_colour = colours.get("Q2", default_colours["Q2"])
    q3_colour = colours.get("Q3", default_colours["Q3"])
    q4_colour = colours.get("Q4", default_colours["Q4"])

    if graph_quadrants:
        try:
            qdata  = json.loads(graph_quadrants)
            q1_flag = str(qdata.get("Q1", True)).lower()
            q2_flag = str(qdata.get("Q2", True)).lower()
            q3_flag = str(qdata.get("Q3", True)).lower()
            q4_flag = str(qdata.get("Q4", True)).lower()
        except (json.JSONDecodeError, ValueError):
            quadrants = settings.get("quadrants", {"Q1": True, "Q2": True, "Q3": True, "Q4": True})
            q1_flag = str(quadrants.get("Q1", True)).lower()
            q2_flag = str(quadrants.get("Q2", True)).lower()
            q3_flag = str(quadrants.get("Q3", True)).lower()
            q4_flag = str(quadrants.get("Q4", True)).lower()
    else:
        quadrants = settings.get("quadrants", {"Q1": True, "Q2": True, "Q3": True, "Q4": True})
        q1_flag = str(quadrants.get("Q1", True)).lower()
        q2_flag = str(quadrants.get("Q2", True)).lower()
        q3_flag = str(quadrants.get("Q3", True)).lower()
        q4_flag = str(quadrants.get("Q4", True)).lower()

    try:
        file_stream = BytesIO(file_bytes)
        file_stream.seek(0)

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_input:
            tmp_input.write(file_stream.read())
            input_path = tmp_input.name

        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_output:
            output_plot_path = tmp_output.name

        plot_title = os.path.splitext(filename)[0]
        subprocess.run(
            [
                "Rscript", r_script,
                input_path, output_plot_path,
                str(include_timestamp).lower(),
                q1_colour, q2_colour, q3_colour, q4_colour,
                plot_title,
                q1_flag, q2_flag, q3_flag, q4_flag
            ],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )

        with open(output_plot_path, "rb") as f:
            image_bytes = BytesIO(f.read())

        image_bytes.seek(0)
        os.remove(input_path)
        os.remove(output_plot_path)

        png_filename = os.path.splitext(filename)[0] + ".png"
        return send_file(
            image_bytes,
            mimetype="image/png",
            as_attachment=True,
            download_name=png_filename
        )

    except subprocess.CalledProcessError as e:
        flash(f"R script failed: {e.stderr}", "danger")
        return redirect(url_for("index"))

    except Exception as e:
        flash(f"Unexpected error: {str(e)}", "danger")
        return redirect(url_for("index"))


@app.route("/summary_plot/<file_id>")
def summary_plot(file_id):
    if file_id not in in_memory_files:
        return "", 404
    info = in_memory_files[file_id]
    if not info.get("plots_ready"):
        return "", 202  # still processing in background
    plot_bytes = info.get("summary_plot")
    if not plot_bytes:
        return "", 404
    return send_file(BytesIO(plot_bytes), mimetype="image/png")


@app.route("/save_quadrants", methods=["POST"])
def save_quadrants():
    quadrants = request.get_json()
    settings = load_settings()
    settings["quadrants"] = quadrants
    save_settings(settings)
    return "Quadrant settings saved", 200


@app.route("/get_settings")
def get_settings():
    settings = load_settings()
    return jsonify(settings)

@app.route("/get_template_dilutions")
def get_template_dilutions():
    try:
        template_path = load_template_path()
        wb = load_workbook(template_path, data_only=True)
        ws = wb.active
        
        dilutions = []
        for row in range(5, 13):
            cell_value = ws[f'A{row}'].value
            try:
                num_val = float(cell_value)
                if num_val == 0:
                    dilutions.append("0")
                elif num_val >= 1000:
                    dilutions.append(str(int(num_val)))
                elif num_val == int(num_val):
                    dilutions.append(str(int(num_val)))
                else:
                    dilutions.append(str(num_val))
            except (ValueError, TypeError):
                dilutions.append(str(cell_value) if cell_value else "—")
        
        return jsonify({"dilutions": dilutions, "status": "success"})
    
    except FileNotFoundError:
        return jsonify({"dilutions": [], "status": "error", "message": "No template selected"})
    except Exception as e:
        return jsonify({"dilutions": [], "status": "error", "message": str(e)})


@app.route("/create_template_variant", methods=["POST"])
def create_template_variant():
    """
    Duplicate the current template, replace A5:A12 with a new dilution series,
    and save the result as a new named template. The source is never modified.
    """
    data = request.get_json()
    template_name = (data.get("template_name") or "").strip()
    dilutions_raw = data.get("dilutions", [])

    if not template_name:
        return jsonify({"status": "error", "message": "Template name is required."}), 400
    if len(dilutions_raw) != 8:
        return jsonify({"status": "error", "message": "Exactly 8 dilution values are required (one per row A5:A12)."}), 400

    try:
        dilutions = [float(str(d).replace(",", "")) for d in dilutions_raw]
    except (ValueError, TypeError):
        return jsonify({"status": "error", "message": "All dilution values must be numbers."}), 400

    try:
        source_path = load_template_path()
    except FileNotFoundError:
        return jsonify({"status": "error", "message": "No source template is currently selected."}), 400

    safe_name = re.sub(r"[^\w\s-]", "", template_name).strip().replace(" ", "_")
    if not safe_name:
        return jsonify({"status": "error", "message": "Template name contains no valid characters."}), 400

    os.makedirs("excel_templates", exist_ok=True)
    output_path = os.path.join("excel_templates", f"{safe_name}.xlsx")

    if os.path.exists(output_path):
        return jsonify({"status": "error", "message": f"A file named '{safe_name}.xlsx' already exists. Choose a different name."}), 400

    try:
        import openpyxl as _xl
        wb = _xl.load_workbook(source_path)
        ws = wb.active
        for i, val in enumerate(dilutions):
            ws[f"A{5 + i}"] = val
        wb.save(output_path)
    except Exception as e:
        return jsonify({"status": "error", "message": f"Failed to write template: {e}"}), 500

    # Register in settings so it appears in the dropdown
    settings = load_settings()
    if "custom_templates" not in settings:
        settings["custom_templates"] = {}
    settings["custom_templates"][template_name] = output_path
    save_settings(settings)

    return jsonify({"status": "success", "name": template_name, "path": output_path})


@app.route("/delete_custom_template", methods=["POST"])
def delete_custom_template():
    """Remove a user-created template from settings (optionally deletes the file too)."""
    data = request.get_json()
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"status": "error", "message": "No name provided."}), 400

    settings = load_settings()
    custom = settings.get("custom_templates", {})
    if name not in custom:
        return jsonify({"status": "error", "message": "Template not found."}), 404

    path = custom.pop(name)
    save_settings(settings)

    # Delete the file if it's in excel_templates/ and still exists
    if path.startswith("excel_templates/") and os.path.exists(path):
        try:
            os.remove(path)
        except OSError:
            pass

    return jsonify({"status": "success"})


@app.route("/validate_csv_mode", methods=["POST"])
def validate_csv_mode_route():
    """JSON API: check whether the uploaded CSV matches the selected data mode."""
    file = request.files.get("csv_file")
    selected_mode = request.form.get("data_mode", "standard")

    if not file:
        return jsonify({"ok": True})

    csv_bytes = BytesIO(file.read())
    csv_bytes.seek(0)

    ok, detected, message = validate_csv_mode(csv_bytes, selected_mode)
    return jsonify({
        "ok": ok,
        "detected_mode": detected,
        "message": message,
    })


@app.route("/save_preset", methods=["POST"])
def save_preset():
    data = request.get_json()
    settings = load_settings()
    settings["presets"][data["name"]] = data["colours"]
    save_settings(settings)
    return jsonify({"status": "success", "message": "Preset saved."}), 200


@app.route("/delete_preset", methods=["POST"])
def delete_preset():
    data = request.get_json()
    settings = load_settings()
    settings["presets"].pop(data["name"], None)
    save_settings(settings)
    return jsonify({"status": "success", "message": "Preset deleted."}), 200


@app.route("/set_active_preset", methods=["POST"])
def set_active_preset():
    data = request.get_json()
    name = data.get("name")
    if not name:
        return "No preset name provided", 400

    settings = load_settings()
    if "presets" not in settings or name not in settings["presets"]:
        return "Preset not found", 404

    settings["selected_preset"] = name
    save_settings(settings)

    return "Preset updated", 200


def _run_fitting(file_id, include_lod_override=None):
    """
    Run fit_sigmoids.R for the given file_id and store results in memory.
    include_lod_override: True/False to override settings, None to use settings.
    Returns a Flask response.
    """
    file_info = in_memory_files[file_id]
    file_bytes = file_info["data"]
    filename = file_info["name"]

    try:
        file_stream = BytesIO(file_bytes)
        file_stream.seek(0)

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_excel:
            tmp_excel.write(file_stream.read())
            excel_path = tmp_excel.name

        output_dir = tempfile.mkdtemp(prefix="sigmoid_")

        sigmoid_csv_path = os.path.join(output_dir, "sigmoidData.csv")
        generate_sigmoid_csv(excel_path, sigmoid_csv_path)

        settings = load_settings()
        include_timestamp = settings.get("timestamp_in_filename", True)

        assay_title = os.path.splitext(filename)[0]
        assay_title = re.sub(r'_\d{4}-\d{2}-\d{2}$', '', assay_title)

        if include_timestamp:
            timestamp = datetime.now().strftime("%Y-%m-%d")
        else:
            timestamp = ""

        r_script = os.path.join(os.getcwd(), "fit_sigmoids.R")
        r2_threshold = str(settings.get("sigmoid_r2_threshold", 0.5))

        if include_lod_override is not None:
            lod_bool = include_lod_override
        else:
            lod_bool = settings.get("lod_censor_include", False)
        include_lod = "TRUE" if lod_bool else "FALSE"

        _proc_start = time.time()
        logger.info("FITTING  starting fit_sigmoids.R \u2026")
        subprocess.run(
            ["Rscript", r_script, sigmoid_csv_path, output_dir, assay_title, timestamp, r2_threshold, include_lod],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        logger.info("FITTING  R complete in %.1fs", time.time() - _proc_start)

        if assay_title and timestamp:
            ic50_filename = f"IC50s_{assay_title}_{timestamp}.csv"
        elif assay_title:
            ic50_filename = f"IC50s_{assay_title}.csv"
        elif timestamp:
            ic50_filename = f"IC50s_{timestamp}.csv"
        else:
            ic50_filename = "IC50s.csv"

        ic50_path = os.path.join(output_dir, ic50_filename)
        plot_files = [f for f in os.listdir(output_dir) if f.endswith('.png')]

        output_files = {}
        if os.path.exists(ic50_path):
            with open(ic50_path, 'rb') as f:
                output_files[ic50_filename] = f.read()
        if os.path.exists(sigmoid_csv_path):
            with open(sigmoid_csv_path, 'rb') as f:
                output_files['sigmoidData.csv'] = f.read()
        for plot_file in plot_files:
            plot_path = os.path.join(output_dir, plot_file)
            with open(plot_path, 'rb') as f:
                output_files[plot_file] = f.read()

        os.remove(excel_path)
        for f in os.listdir(output_dir):
            os.remove(os.path.join(output_dir, f))
        os.rmdir(output_dir)

        fitting_id = uuid.uuid4().hex
        in_memory_files[fitting_id] = {
            "data": output_files,
            "name": "sigmoid_fitting_results",
            "type": "sigmoid_results",
            "ic50_filename": ic50_filename,
            "excel_file_id": file_id,
            "include_lod": lod_bool,
        }
        # Store fitting_id server-side so analysis hub can unlock comparison card
        # Also clear any cached comparison since the IC50s have changed
        in_memory_files[file_id]["fitting_id"] = fitting_id
        in_memory_files[file_id].pop("comparison_id", None)

        _proc_elapsed = round(time.time() - _proc_start, 1)
        return render_template(
            "curve_fitting_results.html",
            fitting_id=fitting_id,
            excel_file_id=file_id,
            output_files=list(output_files.keys()),
            ic50_filename=ic50_filename,
            lod_used=lod_bool,
            settings=load_settings(),
            processing_time=_proc_elapsed,
        )

    except subprocess.CalledProcessError as e:
        flash(f"R script failed: {e.stderr}", "danger")
        return redirect(url_for("analysis_hub", file_id=file_id))
    except Exception as e:
        flash(f"Curve fitting error: {str(e)}", "danger")
        return redirect(url_for("analysis_hub", file_id=file_id))


@app.route("/perform_curve_fitting", methods=["POST"])
def perform_curve_fitting():
    file_id = request.form.get("file_id")
    if not file_id or file_id not in in_memory_files:
        flash("No Excel file found for curve fitting.", "danger")
        return redirect(url_for("index"))

    file_info = in_memory_files[file_id]

    # Return cached results immediately if already computed for this file
    existing_fitting_id = file_info.get("fitting_id")
    if existing_fitting_id and existing_fitting_id in in_memory_files:
        return redirect(url_for("curve_fitting_results", fitting_id=existing_fitting_id))

    return _run_fitting(file_id)


@app.route("/refit_sigmoids", methods=["POST"])
def refit_sigmoids():
    """Re-run fitting with a toggled LOD setting, bypassing the cache."""
    file_id = request.form.get("file_id")
    if not file_id or file_id not in in_memory_files:
        flash("No Excel file found for curve fitting.", "danger")
        return redirect(url_for("index"))

    raw = request.form.get("include_lod", "false").strip().lower()
    include_lod_override = raw == "true"

    return _run_fitting(file_id, include_lod_override=include_lod_override)

@app.route("/download_sigmoid/<fitting_id>/<filename>")
def download_sigmoid(fitting_id, filename):
    if fitting_id not in in_memory_files:
        flash("Fitting results not found.", "danger")
        return redirect(url_for("index"))

    file_info = in_memory_files[fitting_id]
    if file_info.get("type") != "sigmoid_results":
        flash("Invalid file type.", "danger")
        return redirect(url_for("index"))

    if filename not in file_info["data"]:
        flash("File not found.", "danger")
        return redirect(url_for("index"))

    file_bytes = BytesIO(file_info["data"][filename])
    file_bytes.seek(0)

    if filename.endswith('.csv'):
        mimetype = 'text/csv'
    elif filename.endswith('.png'):
        mimetype = 'image/png'
    else:
        mimetype = 'application/octet-stream'

    download_name = filename
    settings = load_settings()
    if settings.get("timestamp_in_filename", False):
        ts = datetime.now().strftime("%Y-%m-%d")
        base, ext = os.path.splitext(filename)
        # Skip if filename already contains a date stamp (e.g. IC50 files)
        if not re.search(r'\d{4}-\d{2}-\d{2}', base):
            download_name = f"{base}_{ts}{ext}"

    return send_file(
        file_bytes,
        as_attachment=True,
        download_name=download_name,
        mimetype=mimetype
    )


@app.route("/cached_sigmoid_graph/<fitting_id>")
def cached_sigmoid_graph(fitting_id):
    """Serve the last-generated sigmoid PNG from in-memory cache (no R re-run)."""
    if fitting_id not in in_memory_files:
        return "", 404
    png_bytes = in_memory_files[fitting_id].get("data", {}).get("sigmoid_combined.png")
    if not png_bytes:
        return "", 404
    return send_file(BytesIO(png_bytes), mimetype="image/png")


@app.route("/generate_sigmoid_graph/<fitting_id>")
def generate_sigmoid_graph(fitting_id):
    if fitting_id not in in_memory_files:
        return jsonify({"error": "Fitting results not found"}), 404

    file_info = in_memory_files[fitting_id]
    if file_info.get("type") != "sigmoid_results":
        return jsonify({"error": "Invalid type"}), 400

    show_good     = request.args.get("good",     "true").lower() == "true"
    show_unstable = request.args.get("unstable", "true").lower() == "true"
    show_poor_fit = request.args.get("poor_fit", "true").lower() == "true"
    # show_lod: page toggle overrides settings when explicitly provided
    _lod_param = request.args.get("show_lod", None)
    if _lod_param is not None:
        show_lod_bool = _lod_param.lower() == "true"
    else:
        _plot_settings_lod = load_settings()
        show_lod_bool = _plot_settings_lod.get("lod_censor_include", False)

    tmp_dir = None
    try:
        tmp_dir = tempfile.mkdtemp(prefix="sigplot_")

        raw_csv = os.path.join(tmp_dir, "sigmoidData.csv")
        with open(raw_csv, "wb") as f:
            f.write(file_info["data"]["sigmoidData.csv"])

        ic50_filename = file_info["ic50_filename"]
        ic50_csv = os.path.join(tmp_dir, ic50_filename)
        with open(ic50_csv, "wb") as f:
            f.write(file_info["data"][ic50_filename])

        output_png = os.path.join(tmp_dir, "sigmoid_combined.png")

        r_script = os.path.join(os.getcwd(), "plot_sigmoids.R")
        subprocess.run(
            ["Rscript", r_script, raw_csv, ic50_csv, output_png,
             str(show_good).lower(), str(show_unstable).lower(),
             str(show_lod_bool).lower(), str(show_poor_fit).lower()],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )

        with open(output_png, "rb") as f:
            png_bytes = f.read()

        # Store latest render for download
        in_memory_files[fitting_id]["data"]["sigmoid_combined.png"] = png_bytes

        return send_file(BytesIO(png_bytes), mimetype="image/png")

    except subprocess.CalledProcessError as e:
        return jsonify({"error": f"R script failed: {e.stderr}"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if tmp_dir and os.path.exists(tmp_dir):
            for fn in os.listdir(tmp_dir):
                try:
                    os.remove(os.path.join(tmp_dir, fn))
                except Exception:
                    pass
            try:
                os.rmdir(tmp_dir)
            except Exception:
                pass


# ════════════════════════════════════════════════════════════════
# Titre Comparison (refactored into reusable helper)
# ════════════════════════════════════════════════════════════════

@app.route("/compare_titres", methods=["POST"])
def compare_titres():
    """Legacy POST route — still works for form submissions."""
    excel_file_id = request.form.get("excel_file_id")
    fitting_id = request.form.get("fitting_id")
    
    if not excel_file_id or excel_file_id not in in_memory_files:
        flash("Excel results not found.", "danger")
        return redirect(url_for("index"))
    
    if not fitting_id or fitting_id not in in_memory_files:
        flash("Curve fitting results not found. Please perform curve fitting first.", "danger")
        return redirect(url_for("index"))

    return _run_comparison(excel_file_id, fitting_id)


def _run_comparison(excel_file_id, fitting_id):
    """
    Shared comparison logic used by both POST and GET routes.
    
    R now handles both the NT50 linear interpolation (reading Plate sheets
    directly from the Excel file) and the IC50 comparison, so Python just
    orchestrates temp files and reads back the CSV results.
    """
    try:
        excel_info = in_memory_files[excel_file_id]
        excel_bytes = BytesIO(excel_info["data"])

        fitting_info = in_memory_files[fitting_id]
        if fitting_info.get("type") != "sigmoid_results":
            flash("Invalid fitting results.", "danger")
            return redirect(url_for("index"))

        ic50_filename = fitting_info.get("ic50_filename", "IC50s.csv")
        if ic50_filename not in fitting_info["data"]:
            flash("IC50 file not found in fitting results.", "danger")
            return redirect(url_for("index"))

        ic50_bytes = fitting_info["data"][ic50_filename]

        # Write Excel and IC50 CSV to temp files for R
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_excel:
            tmp_excel.write(excel_bytes.getvalue())
            excel_path = tmp_excel.name

        with tempfile.NamedTemporaryFile(delete=False, suffix=".csv") as tmp_ic50:
            tmp_ic50.write(ic50_bytes)
            ic50_path = tmp_ic50.name

        output_dir = tempfile.mkdtemp(prefix="comparison_")

        # R now receives the Excel file directly (not a pre-computed NT50 CSV)
        r_script = os.path.join(os.getcwd(), "compare_titres.R")
        cmp_settings = load_settings()
        disagreement_threshold = str(cmp_settings.get("comparison_disagreement_threshold", 1.0))

        _proc_start = time.time()
        logger.info("COMPARE  starting compare_titres.R \u2026")
        result = subprocess.run(
            ["Rscript", r_script, excel_path, ic50_path, output_dir, disagreement_threshold],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=180,
        )
        logger.info("COMPARE  R complete in %.1fs", time.time() - _proc_start)

        if result.stderr:
            logger.warning("compare_titres.R stderr:\n%s", result.stderr.strip())

        # Collect output files
        output_files = {}
        for fname in ['comparison_stats.csv', 'merged_titres.csv',
                       'top_mismatches.csv', 'titre_comparison.png',
                       'titre_comparison_interactive.html']:
            filepath = os.path.join(output_dir, fname)
            if os.path.exists(filepath):
                with open(filepath, 'rb') as f:
                    output_files[fname] = f.read()

        # Clean up temp files
        os.remove(excel_path)
        os.remove(ic50_path)
        for f in os.listdir(output_dir):
            os.remove(os.path.join(output_dir, f))
        os.rmdir(output_dir)

        # Store results in memory
        comparison_id = uuid.uuid4().hex
        in_memory_files[comparison_id] = {
            "data": output_files,
            "name": "titre_comparison",
            "type": "comparison_results"
        }

        # Parse stats and mismatches for template rendering
        import csv as py_csv

        stats = {}
        if 'comparison_stats.csv' in output_files:
            csv_data = output_files['comparison_stats.csv'].decode('utf-8')
            reader = py_csv.DictReader(csv_data.splitlines())
            stats = next(reader)

        mismatches = []
        if 'merged_titres.csv' in output_files:
            csv_data = output_files['merged_titres.csv'].decode('utf-8')
            reader = py_csv.DictReader(csv_data.splitlines())
            all_rows = list(reader)
            for row in all_rows:
                try:
                    lfd = float(row.get('log2_fold_difference', 0) or 0)
                except ValueError:
                    lfd = 0
                if abs(lfd) >= 2:
                    mismatches.append({
                        'Sample':               row.get('Sample_ID', ''),
                        'Virus':                row.get('Pseudotype', ''),
                        'NT50':                 row.get('NT50 (Linear Interpolation)', ''),
                        'IC50_Titre':           row.get('NT50 / IC50 (Curve Fitting)', ''),
                        'Log2_Fold_Difference': lfd,
                        'Quality':              row.get('Sigmoid Quality', ''),
                    })
            mismatches.sort(key=lambda r: abs(r['Log2_Fold_Difference']), reverse=True)

        # Cache parsed render data so future visits skip reprocessing
        has_plot = 'titre_comparison.png' in output_files
        in_memory_files[comparison_id]["stats"]      = stats
        in_memory_files[comparison_id]["mismatches"] = mismatches
        in_memory_files[comparison_id]["has_plot"]   = has_plot
        # Store back-reference so the hub and compare_titres_page can find the cache
        in_memory_files[excel_file_id]["comparison_id"] = comparison_id

        _proc_elapsed = round(time.time() - _proc_start, 1)
        return render_template(
            "titre_comparison_results.html",
            comparison_id=comparison_id,
            excel_file_id=excel_file_id,
            stats=stats,
            mismatches=mismatches,
            has_plot=has_plot,
            settings=load_settings(),
            processing_time=_proc_elapsed,
        )

    except subprocess.TimeoutExpired:
        flash("Comparison timed out (>3 min). Check your terminal for R output.", "danger")
        return redirect(url_for("analysis_hub", file_id=excel_file_id))
    except subprocess.CalledProcessError as e:
        logger.error("compare_titres.R FAILED\nSTDOUT: %s\nSTDERR: %s", e.stdout, e.stderr)
        flash(f"R script failed — see terminal for details. Error: {(e.stderr or e.stdout or '').strip()[:300]}", "danger")
        return redirect(url_for("analysis_hub", file_id=excel_file_id))
    except Exception as e:
        logger.exception("_run_comparison exception: %s", e)
        flash(f"Comparison error: {str(e)}", "danger")
        return redirect(url_for("analysis_hub", file_id=excel_file_id))


@app.route("/download_comparison/<comparison_id>/<filename>")
def download_comparison(comparison_id, filename):
    if comparison_id not in in_memory_files:
        flash("Comparison results not found.", "danger")
        return redirect(url_for("index"))
    
    file_info = in_memory_files[comparison_id]
    if file_info.get("type") != "comparison_results":
        flash("Invalid file type.", "danger")
        return redirect(url_for("index"))
    
    if filename not in file_info["data"]:
        flash("File not found.", "danger")
        return redirect(url_for("index"))
    
    file_bytes = BytesIO(file_info["data"][filename])
    file_bytes.seek(0)
    
    if filename.endswith('.csv'):
        mimetype = 'text/csv'
        as_attachment = True
    elif filename.endswith('.png'):
        mimetype = 'image/png'
        as_attachment = True
    elif filename.endswith('.html'):
        mimetype = 'text/html'
        as_attachment = request.args.get('download') == '1'
    else:
        mimetype = 'application/octet-stream'
        as_attachment = True
    
    download_name = filename
    if as_attachment:
        settings = load_settings()
        if settings.get("timestamp_in_filename", False):
            ts = datetime.now().strftime("%Y-%m-%d")
            base, ext = os.path.splitext(filename)
            download_name = f"{base}_{ts}{ext}"

    return send_file(
        file_bytes,
        as_attachment=as_attachment,
        download_name=download_name if as_attachment else None,
        mimetype=mimetype
    )


# ════════════════════════════════════════════════════════════════
# LEGACY: Keep /results/<file_id> route as redirect to hub
# ════════════════════════════════════════════════════════════════

@app.route("/results/<file_id>")
def view_results(file_id):
    """Legacy route — redirects to Data Analysis."""
    return redirect(url_for("analysis_hub", file_id=file_id))


# ════════════════════════════════════════════════════════════════
# PLATE MAPPER
# ════════════════════════════════════════════════════════════════

PM_PRESETS_FILE = os.path.join(os.path.dirname(__file__), 'plate_mapper_presets.json')


def pm_load_presets():
    if os.path.exists(PM_PRESETS_FILE):
        with open(PM_PRESETS_FILE, 'r') as f:
            return json.load(f)
    return {}


def _pm_write_presets(presets):
    with open(PM_PRESETS_FILE, 'w') as f:
        json.dump(presets, f, indent=4)


def pm_save_preset(name, settings):
    presets = pm_load_presets()
    presets[name] = settings
    _pm_write_presets(presets)
    return True


def pm_delete_preset(name):
    presets = pm_load_presets()
    if name in presets:
        del presets[name]
        _pm_write_presets(presets)
        return True
    return False


def pm_format_id(entry, prefix, suffix='', id_length=0, pad_char='0'):
    if not entry or entry.strip() == "":
        return None
    if entry.startswith('!'):
        return entry[1:]
    if id_length > 0:
        entry = entry.zfill(id_length) if pad_char == '0' else entry.rjust(id_length, pad_char)
    return f"{prefix}{entry}{suffix}"


def pm_build_grid(sample_ids, replicate_count, layout_mode,
                  pos_label='PosCtrl', neg_label='NegCtrl', ctrl3_label='Ctrl3',
                  include_pos=True, include_neg=True, include_ctrl3=True):
    grid = [["" for _ in range(12)] for _ in range(8)]
    if layout_mode == 'vertical':
        fill_positions = []
        control_labels = []
        if include_pos:
            control_labels += [pos_label] * replicate_count
        if include_neg:
            control_labels += [neg_label] * replicate_count
        if include_ctrl3:
            control_labels += [ctrl3_label] * replicate_count
        control_col = 0
        control_row = 0
        for label in control_labels:
            if control_row >= 8:
                control_row = 0
                control_col += 1
                if control_col >= 12:
                    break
            grid[control_row][control_col] = label
            control_row += 1
        for col in range(12):
            for row in range(8):
                if grid[row][col] == "":
                    fill_positions.append((row, col))
        idx = 0
        for sid in sample_ids:
            for _ in range(replicate_count):
                if idx >= len(fill_positions):
                    break
                r, c = fill_positions[idx]
                grid[r][c] = sid
                idx += 1
    elif layout_mode == 'horizontal':
        current_row = 0
        col = 0
        if include_pos:
            for rep in range(replicate_count):
                grid[current_row][col + rep] = pos_label
            current_row += 1
        if include_neg:
            for rep in range(replicate_count):
                grid[current_row][col + rep] = neg_label
            current_row += 1
        if include_ctrl3:
            for rep in range(replicate_count):
                grid[current_row][col + rep] = ctrl3_label
            current_row += 1
        sample_idx = 0
        while sample_idx < len(sample_ids):
            if current_row >= 8:
                current_row = 0
                col += replicate_count
                if col + replicate_count > 12:
                    break
            for rep in range(replicate_count):
                if col + rep < 12:
                    grid[current_row][col + rep] = sample_ids[sample_idx]
            current_row += 1
            sample_idx += 1
    return grid


@app.route('/plate_mapper', methods=['GET', 'POST'])
def plate_mapper():
    if request.method == 'POST':
        prefix = request.form.get('prefix', '')
        suffix = request.form.get('suffix', '')
        pos_label = request.form.get('pos_label', '').strip() or "Control 1"
        neg_label = request.form.get('neg_label', '').strip() or "Control 2"
        ctrl3_label = request.form.get('ctrl3_label', '').strip() or "Control 3"
        include_pos = 'include_pos' in request.form
        include_neg = 'include_neg' in request.form
        include_ctrl3 = 'include_ctrl3' in request.form
        replicate_count = int(request.form.get('replicate_count', 2) or 2)
        id_length = int(request.form.get('id_length', 0) or 0)
        pad_char = (request.form.get('pad_char', '0') or '0')[:1] or '0'
        layout_mode = request.form.get('layout_mode', 'vertical')
        plate_count = int(request.form.get('plate_count', 1) or 1)

        plate_labels = []
        for i in range(1, plate_count + 1):
            label = request.form.get(f'plate_label_{i}', '').strip()
            plate_labels.append(label or f"Plate {i}")

        plates = []
        for i in range(1, plate_count + 1):
            raw = request.form.get(f'plate_{i}', '')
            nums = [s.strip() for s in raw.replace(';', ',').split(',') if s.strip()]
            formatted = [
                pm_format_id(n, prefix, suffix, id_length, pad_char)
                for n in nums
                if pm_format_id(n, prefix, suffix, id_length, pad_char)
            ]
            plates.append(pm_build_grid(
                formatted, replicate_count, layout_mode,
                pos_label, neg_label, ctrl3_label,
                include_pos, include_neg, include_ctrl3
            ))

        control_labels = set()
        if include_pos:
            control_labels.add(pos_label)
        if include_neg:
            control_labels.add(neg_label)
        if include_ctrl3:
            control_labels.add(ctrl3_label)

        return render_template('plate_mapper_results.html', plates=plates, plate_labels=plate_labels,
                               control_labels=control_labels)
    return render_template('plate_mapper.html')


@app.route('/plate_mapper/settings')
def plate_mapper_settings():
    return render_template('plate_mapper_settings.html')


@app.route('/plate_mapper/api/presets', methods=['GET'])
def pm_api_get_presets():
    return jsonify(pm_load_presets())


@app.route('/plate_mapper/api/presets', methods=['POST'])
def pm_api_save_preset():
    data = request.json
    if not data or "name" not in data or "settings" not in data:
        return jsonify({"error": "Invalid data"}), 400
    if pm_save_preset(data["name"], data["settings"]):
        return jsonify({"message": f'Preset "{data["name"]}" saved.'})
    return jsonify({"error": "Failed to save preset"}), 500


@app.route('/plate_mapper/api/presets/<name>', methods=['DELETE'])
def pm_api_delete_preset(name):
    if pm_delete_preset(name):
        return jsonify({"message": f'Preset "{name}" deleted.'})
    return jsonify({"error": "Preset not found"}), 404


# ════════════════════════════════════════════════════════════════
# ELISA PROCESSOR
# ════════════════════════════════════════════════════════════════

_ELISA_BASE_FONT = Font(name='Aptos Narrow', size=12)
_ELISA_BOLD_FONT = Font(name='Aptos Narrow', size=12, bold=True)
_ELISA_THIN   = Side(border_style='thin')
_ELISA_MEDIUM = Side(border_style='medium')
_ELISA_NONE   = Side(border_style=None)
_ELISA_CENTER  = Alignment(horizontal='center', vertical='center')
_ELISA_LEFT    = Alignment(horizontal='left',   vertical='center')
_ELISA_VCENTER = Alignment(vertical='center')


def _elisa_border(left=False, right=False, top=False, bottom=False,
                  med_right=False, med_bottom=False):
    return Border(
        left   = _ELISA_THIN   if left        else _ELISA_NONE,
        right  = _ELISA_MEDIUM if med_right   else (_ELISA_THIN if right  else _ELISA_NONE),
        top    = _ELISA_THIN   if top         else _ELISA_NONE,
        bottom = _ELISA_MEDIUM if med_bottom  else (_ELISA_THIN if bottom else _ELISA_NONE),
    )


def _elisa_set(ws, row, col, value=None, font=None, align=None, border=None, fmt=None):
    c = ws.cell(row=row, column=col)
    c.value     = value
    c.font      = font or _ELISA_BASE_FONT
    c.alignment = align if align is not None else _ELISA_VCENTER
    if border is not None: c.border       = border
    if fmt    is not None: c.number_format = fmt
    return c


_ELISA_ROW_LBLS_LOWER = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h']
_ELISA_ROW_LBLS_UPPER = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']


def _elisa_populate_sheet(ws, n_proteins, protein_names, date_str, sample_grid,
                           protein_grids, sera_dilution='1 IN 400', notes=''):
    _elisa_set(ws, 1, 1, 'Date')
    _elisa_set(ws, 1, 2, date_str)
    _elisa_set(ws, 1, 3, 'ELISA')
    sera_text = f'SERA DILUTED {sera_dilution.strip()}' if sera_dilution else 'SERA DILUTED'
    _elisa_set(ws, 1, 4, sera_text, align=_ELISA_LEFT)
    ws.merge_cells('D1:M1')

    _elisa_set(ws, 2, 1, 'Enter your sample names here (duplicate)')

    c = ws.cell(row=3, column=1)
    c.font   = _ELISA_BASE_FONT
    c.border = _elisa_border(left=True, top=True)

    for ci in range(12):
        col     = 2 + ci
        is_last = ci == 11
        _elisa_set(ws, 3, col, ci + 1, align=_ELISA_CENTER,
                   border=_elisa_border(right=is_last, top=True))

    for ri, lbl in enumerate(_ELISA_ROW_LBLS_LOWER):
        row     = 4 + ri
        is_last = ri == 7
        _elisa_set(ws, row, 1, lbl, border=_elisa_border(left=True, bottom=is_last))

    for ri in range(8):
        for ci in range(12):
            row      = 4 + ri
            col      = 2 + ci
            v        = sample_grid[ri][ci]
            last_row = ri == 7
            last_col = ci == 11
            c = ws.cell(row=row, column=col, value=v if v else None)
            c.font      = _ELISA_BASE_FONT
            c.alignment = _ELISA_CENTER
            if last_row or last_col:
                c.border = _elisa_border(med_right=last_col, med_bottom=last_row)

    plate_base = 15
    for k in range(n_proteins):
        lbl_row    = plate_base + k * 12
        hdr_row    = lbl_row + 1
        data_start = lbl_row + 2

        _elisa_set(ws, lbl_row, 1, protein_names[k])

        c = ws.cell(row=hdr_row, column=1)
        c.font   = _ELISA_BASE_FONT
        c.border = _elisa_border(left=True, top=True)
        for ci in range(12):
            col     = 2 + ci
            is_last = ci == 11
            _elisa_set(ws, hdr_row, col, ci + 1, align=_ELISA_CENTER,
                       border=_elisa_border(right=is_last, top=True))

        for ri, lbl in enumerate(_ELISA_ROW_LBLS_UPPER):
            row     = data_start + ri
            is_last = ri == 7
            _elisa_set(ws, row, 1, lbl, border=_elisa_border(left=True, bottom=is_last))

        for ri in range(8):
            for ci in range(12):
                row = data_start + ri
                col = 2 + ci
                v   = protein_grids[k][ri][ci]
                if v not in (None, ''):
                    try:
                        v = float(v)
                    except (TypeError, ValueError):
                        pass
                else:
                    v = None
                c = ws.cell(row=row, column=col, value=v)
                c.font          = _ELISA_BASE_FONT
                c.alignment     = _ELISA_CENTER
                c.number_format = '0.0000'

    _elisa_set(ws, 2, 16, 'SampleID', font=_ELISA_BOLD_FONT, align=_ELISA_CENTER)
    for k in range(n_proteins):
        prot_lbl_row = plate_base + k * 12
        _elisa_set(ws, 2, 17 + k, f'=A{prot_lbl_row}', font=_ELISA_BOLD_FONT, align=_ELISA_CENTER)

    out_row = 3
    for plate_col in range(12):
        col_ltr = get_column_letter(2 + plate_col)
        for sample_row in range(8):
            src = 4 + sample_row
            c   = ws.cell(row=out_row, column=16, value=f'={col_ltr}{src}')
            c.font      = _ELISA_BASE_FONT
            c.alignment = _ELISA_CENTER
            out_row += 1

    for k in range(n_proteins):
        ds         = plate_base + k * 12 + 2
        target_col = 17 + k
        out_row    = 3
        for plate_col in range(12):
            col_ltr = get_column_letter(2 + plate_col)
            for sample_row in range(8):
                src = ds + sample_row
                c   = ws.cell(row=out_row, column=target_col,
                              value=f'={col_ltr}{src}')
                c.font          = _ELISA_BASE_FONT
                c.alignment     = _ELISA_CENTER
                c.number_format = '0.0000'
                out_row += 1

    STAT_HDRS = ['Pmean', 'Pstdv', 'Nmean', 'Nstdv', 'Pcv', 'Ncv']
    for k in range(n_proteins):
        hdr_row = 3 + k * 3
        f_row   = hdr_row + 1
        ds      = plate_base + k * 12 + 2
        pn      = protein_names[k]
        pa      = ds

        for i, suffix in enumerate(STAT_HDRS):
            col = 22 + i
            _elisa_set(ws, hdr_row, col, f'{pn}_{suffix}',
                       border=_elisa_border(left=(i == 0), right=(i == 5), top=True))

        formulas = [
            f'=AVERAGE(B{pa}:B{pa+1})',
            f'=STDEV(B{pa}:B{pa+1})',
            f'=AVERAGE(B{pa+2}:B{pa+3})',
            f'=STDEV(B{pa+2}:B{pa+3})',
            f'=((W{f_row}/V{f_row})*100)',
            f'=((Y{f_row}/X{f_row})*100)',
        ]
        for i, formula in enumerate(formulas):
            col = 22 + i
            _elisa_set(ws, f_row, col, formula,
                       border=_elisa_border(left=(i == 0), right=(i == 5), bottom=True))

    _YES_FILL = PatternFill(start_color='C8F5DA', end_color='C8F5DA', fill_type='solid')
    _NO_FILL  = PatternFill(start_color='FCD5D5', end_color='FCD5D5', fill_type='solid')
    _WRAP_L   = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for k in range(n_proteins):
        lbl_row   = 17 + k * 8
        pos_start = lbl_row + 1
        pos_end   = lbl_row + 3
        neg_start = lbl_row + 4
        neg_end   = lbl_row + 6
        f_row     = 4 + k * 3
        pn        = protein_names[k]

        ws.merge_cells(start_row=lbl_row, start_column=22,
                       end_row=lbl_row,   end_column=23)
        c = ws.cell(row=lbl_row, column=22)
        c.value, c.font, c.alignment = f'{pn} Plate validation', _ELISA_BOLD_FONT, _ELISA_CENTER
        c.border = _elisa_border(left=True, top=True)
        ws.cell(row=lbl_row, column=23).border = _elisa_border(right=True, top=True)

        ws.merge_cells(start_row=pos_start, start_column=22,
                       end_row=pos_end,     end_column=22)
        ws.merge_cells(start_row=pos_start, start_column=23,
                       end_row=pos_end,     end_column=23)
        c22 = ws.cell(row=pos_start, column=22)
        c22.value, c22.font, c22.alignment = 'Positive control duplicates CV < 15%', _ELISA_BASE_FONT, _WRAP_L
        c23 = ws.cell(row=pos_start, column=23)
        c23.value, c23.font, c23.alignment = f'=IF(Z{f_row}<=15,"Yes!","No!")', _ELISA_BASE_FONT, _ELISA_CENTER
        for r in range(pos_start, pos_end + 1):
            ws.cell(row=r, column=22).border = _elisa_border(left=True)
            ws.cell(row=r, column=23).border = _elisa_border(right=True)

        ws.merge_cells(start_row=neg_start, start_column=22,
                       end_row=neg_end,     end_column=22)
        ws.merge_cells(start_row=neg_start, start_column=23,
                       end_row=neg_end,     end_column=23)
        c22 = ws.cell(row=neg_start, column=22)
        c22.value, c22.font, c22.alignment = 'Negative control duplicate CV < 15%', _ELISA_BASE_FONT, _WRAP_L
        c23 = ws.cell(row=neg_start, column=23)
        c23.value, c23.font, c23.alignment = f'=IF(AA{f_row}<=15,"Yes!","No!")', _ELISA_BASE_FONT, _ELISA_CENTER
        for r in range(neg_start, neg_end + 1):
            is_last = r == neg_end
            ws.cell(row=r, column=22).border = _elisa_border(left=True,  bottom=is_last)
            ws.cell(row=r, column=23).border = _elisa_border(right=True, bottom=is_last)

        for rng in [f'W{pos_start}:W{pos_end}', f'W{neg_start}:W{neg_end}']:
            ws.conditional_formatting.add(
                rng, CellIsRule(operator='equal', formula=['"Yes!"'], fill=_YES_FILL))
            ws.conditional_formatting.add(
                rng, CellIsRule(operator='equal', formula=['"No!"'],  fill=_NO_FILL))

    ws.merge_cells('Y35:AA35')
    c = ws.cell(row=35, column=25)
    c.value, c.font, c.alignment = 'Notes', _ELISA_BOLD_FONT, _ELISA_CENTER
    c.border = _elisa_border(left=True, top=True, bottom=True)
    ws.cell(row=35, column=26).font   = _ELISA_BASE_FONT
    ws.cell(row=35, column=26).border = _elisa_border(top=True, bottom=True)
    ws.cell(row=35, column=26).alignment = _ELISA_VCENTER
    ws.cell(row=35, column=27).font   = _ELISA_BASE_FONT
    ws.cell(row=35, column=27).border = _elisa_border(right=True, top=True, bottom=True)
    ws.cell(row=35, column=27).alignment = _ELISA_VCENTER

    _NOTES_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.merge_cells('Y36:AA40')
    ws.cell(row=36, column=25).value     = notes if notes else None
    ws.cell(row=36, column=25).alignment = _NOTES_ALIGN
    for row in range(36, 41):
        for col in [25, 26, 27]:
            first_col = col == 25
            last_col  = col == 27
            first_row = row == 36
            last_row  = row == 40
            c        = ws.cell(row=row, column=col)
            c.font   = _ELISA_BASE_FONT
            c.border = _elisa_border(left=first_col, right=last_col,
                                     top=first_row, bottom=last_row)
            if not first_row or not first_col:
                c.alignment = _ELISA_VCENTER

    ws.merge_cells(start_row=2, start_column=30, end_row=2, end_column=32)
    ws.cell(row=2, column=30).value     = 'AVERAGES'
    ws.cell(row=2, column=30).font      = _ELISA_BOLD_FONT
    ws.cell(row=2, column=30).alignment = _ELISA_CENTER

    _elisa_set(ws, 3, 30, '=P2', font=_ELISA_BOLD_FONT, align=_ELISA_CENTER)
    for k in range(n_proteins):
        _elisa_set(ws, 3, 31 + k, f'={get_column_letter(17 + k)}2', font=_ELISA_BOLD_FONT)

    for plate_col in range(12):
        for pair_idx in range(4):
            p_row   = 3 + plate_col * 8 + pair_idx * 2
            avg_row = 4 + plate_col * 4 + pair_idx
            _elisa_set(ws, avg_row, 30, f'=P{p_row}')
            for k in range(n_proteins):
                src = get_column_letter(17 + k)
                c   = ws.cell(row=avg_row, column=31 + k,
                              value=f'=AVERAGE({src}{p_row}:{src}{p_row + 1})')
                c.font          = _ELISA_BASE_FONT
                c.alignment     = _ELISA_CENTER
                c.number_format = '0.00'

    if n_proteins >= 2:
        chart_anchor_col = 31 + n_proteins + 1
        anchor_letter    = get_column_letter(chart_anchor_col)
        for i, (a, b) in enumerate(combinations(range(n_proteins), 2)):
            chart = ScatterChart()
            chart.title          = f'{protein_names[a]} vs {protein_names[b]}'
            chart.style          = 13
            chart.x_axis.title   = f'{protein_names[a]} (mean absorbance)'
            chart.y_axis.title   = f'{protein_names[b]} (mean absorbance)'
            chart.legend         = None

            xvalues = Reference(ws, min_col=31 + a, min_row=4, max_row=51)
            yvalues = Reference(ws, min_col=31 + b, min_row=4, max_row=51)
            series  = Series(yvalues, xvalues, title=chart.title)
            series.marker               = Marker(symbol='circle', size=6)
            series.graphicalProperties  = GraphicalProperties(noFill=True)
            series.graphicalProperties.line = LineProperties(noFill=True)
            chart.series.append(series)
            chart.width  = 15
            chart.height = 10
            ws.add_chart(chart, f'{anchor_letter}{3 + i * 22}')

    def _green_white_rule():
        return ColorScaleRule(
            start_type='min', start_color='FFFFFF',
            end_type='max',   end_color='63BE7B',
        )

    for k in range(n_proteins):
        ds  = plate_base + k * 12 + 2
        rng = f'B{ds}:M{ds + 7}'
        ws.conditional_formatting.add(rng, _green_white_rule())

    for k in range(n_proteins):
        col_ltr = get_column_letter(17 + k)
        ws.conditional_formatting.add(f'{col_ltr}3:{col_ltr}98', _green_white_rule())

    for k in range(n_proteins):
        col_ltr = get_column_letter(31 + k)
        ws.conditional_formatting.add(f'{col_ltr}4:{col_ltr}51', _green_white_rule())

    ws.column_dimensions['A'].width = 8.83
    for c in range(2, 14):
        ws.column_dimensions[get_column_letter(c)].width = 17.5
    ws.column_dimensions['N'].width  = 8.83
    ws.column_dimensions['P'].width  = 18.67
    ws.column_dimensions['Q'].width  = 12.33
    ws.column_dimensions['R'].width  = 8.83
    ws.column_dimensions['S'].width  = 8.83
    ws.column_dimensions['T'].width  = 8.83
    ws.column_dimensions['U'].width  = 18.5
    ws.column_dimensions['V'].width  = 14.16
    ws.column_dimensions['W'].width  = 17.16
    ws.column_dimensions['X'].width  = 18.16
    ws.column_dimensions['Y'].width  = 16.67
    ws.column_dimensions['Z'].width  = 17.83
    ws.column_dimensions['AA'].width = 8.83
    for c in range(30, 30 + 1 + n_proteins):
        ws.column_dimensions[get_column_letter(c)].width = 14


def _elisa_build_workbook(n_proteins, n_plates, protein_names, date_str,
                           plates_data, sera_dilution='1 IN 400'):
    wb = Workbook()
    for plate_idx in range(n_plates):
        if plate_idx == 0:
            ws = wb.active
            ws.title = f'Plate {plate_idx + 1}'
        else:
            ws = wb.create_sheet(f'Plate {plate_idx + 1}')

        plate   = plates_data[plate_idx] if plate_idx < len(plates_data) else {}
        sg      = plate.get('sampleGrid') or [[''] * 12 for _ in range(8)]
        pgs     = list(plate.get('proteinGrids') or [])[:n_proteins]
        while len(pgs) < n_proteins:
            pgs.append([[''] * 12 for _ in range(8)])
        notes_p = plate.get('notes', '') or ''

        _elisa_populate_sheet(ws, n_proteins, protein_names, date_str, sg, pgs,
                              sera_dilution=sera_dilution, notes=notes_p)
    return wb


def _elisa_run_generate(data):
    n_proteins = max(1, min(4, int(data.get('nProteins', 4))))
    n_plates   = max(1, int(data.get('nPlates', 1)))

    protein_names = list(data.get('proteinNames') or [])[:n_proteins]
    while len(protein_names) < n_proteins:
        protein_names.append(f'Protein {len(protein_names) + 1}')

    date_str      = data.get('date', '') or ''
    sera_dilution = data.get('seraDilution', '1 IN 400') or '1 IN 400'

    plates_data = list(data.get('plates') or [])
    while len(plates_data) < n_plates:
        plates_data.append({
            'sampleGrid':   [[''] * 12 for _ in range(8)],
            'proteinGrids': [[[''] * 12 for _ in range(8)] for _ in range(n_proteins)],
            'notes':        ''
        })
    plates_data = plates_data[:n_plates]
    title = data.get('title', '') or ''

    return _elisa_build_workbook(n_proteins, n_plates, protein_names, date_str,
                                  plates_data, sera_dilution=sera_dilution), date_str, title


@app.route('/elisa')
def elisa_index():
    return render_template('elisa.html')


@app.route('/elisa/generate', methods=['POST'])
def elisa_generate():
    try:
        data = request.get_json(force=True)
        wb, date_str, title = _elisa_run_generate(data)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_date  = date_str.replace('/', '-').replace(' ', '_') or 'untitled'
        safe_title = ''.join(c for c in title if c.isalnum() or c in ' _-').strip().replace(' ', '_')
        base       = safe_title if safe_title else 'ELISA_results'
        return send_file(
            buf,
            as_attachment=True,
            download_name=f'{base}_{safe_date}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/elisa/preview', methods=['POST'])
def elisa_preview():
    _PREVIEW_SAMPLE = [
        ['Kiovig',   'AVP-0003','AVP-0007','AVP-0011','AVP-0015','AVP-0019','AVP-0023','AVP-0027','AVP-0031','AVP-0035','AVP-0039','AVP-0043'],
        ['Kiovig',   'AVP-0003','AVP-0007','AVP-0011','AVP-0015','AVP-0019','AVP-0023','AVP-0027','AVP-0031','AVP-0035','AVP-0039','AVP-0043'],
        ['NSC',      'AVP-0004','AVP-0008','AVP-0012','AVP-0016','AVP-0020','AVP-0024','AVP-0028','AVP-0032','AVP-0036','AVP-0040','AVP-0044'],
        ['NSC',      'AVP-0004','AVP-0008','AVP-0012','AVP-0016','AVP-0020','AVP-0024','AVP-0028','AVP-0032','AVP-0036','AVP-0040','AVP-0044'],
        ['AVP-0001', 'AVP-0005','AVP-0009','AVP-0013','AVP-0017','AVP-0021','AVP-0025','AVP-0029','AVP-0033','AVP-0037','AVP-0041','AVP-0045'],
        ['AVP-0001', 'AVP-0005','AVP-0009','AVP-0013','AVP-0017','AVP-0021','AVP-0025','AVP-0029','AVP-0033','AVP-0037','AVP-0041','AVP-0045'],
        ['AVP-0002', 'AVP-0006','AVP-0010','AVP-0014','AVP-0018','AVP-0022','AVP-0026','AVP-0030','AVP-0034','AVP-0038','AVP-0042','AVP-0046'],
        ['AVP-0002', 'AVP-0006','AVP-0010','AVP-0014','AVP-0018','AVP-0022','AVP-0026','AVP-0030','AVP-0034','AVP-0038','AVP-0042','AVP-0046'],
    ]
    _PREVIEW_PROT1 = [
        [0.0446,0.7930,0.5243,1.0941,0.3528,0.1453,0.3061,0.2359,0.4177,1.7196,0.7080,0.9184],
        [0.0456,0.9159,1.2170,1.1803,0.2652,0.1088,0.3788,0.1819,0.1437,0.3307,0.1275,0.4220],
        [0.0426,0.2927,0.3134,1.0127,0.6571,0.4860,0.3205,0.2492,0.2187,0.2041,0.4240,0.8170],
        [0.0460,0.3296,0.5417,1.9179,0.1457,0.6373,0.2740,0.1581,0.1960,0.1155,1.9015,0.4219],
        [0.4569,0.4267,0.5486,0.2707,0.1640,0.3107,0.3702,0.6452,0.9274,0.6711,0.3325,0.4030],
        [0.6928,0.6168,0.7851,0.5040,0.2369,0.4224,0.2781,0.2343,0.4753,0.4608,0.0457,0.2329],
        [0.7838,0.5156,2.1424,0.4240,0.8384,0.5607,0.4848,0.6179,0.2986,0.3393,0.2703,0.3647],
        [0.7682,0.7724,1.9328,0.2569,0.5608,0.5392,0.4199,0.3748,0.3078,0.4225,0.1853,0.2920],
    ]
    _PREVIEW_PROT2 = [
        [0.0449,1.0723,0.2499,0.1876,0.3006,0.1083,0.1125,0.1763,0.7565,0.1689,0.1834,0.2501],
        [0.0478,0.8030,0.2277,0.1731,0.2020,0.1356,0.2038,0.1844,0.3543,0.4026,0.1123,0.1169],
        [0.0448,0.2358,0.0965,0.2486,0.5252,0.1748,0.1007,0.2368,0.1676,0.1647,0.2897,0.9734],
        [0.0416,0.1823,0.1362,0.2406,0.1730,0.4690,0.1115,0.1049,0.1448,0.0731,1.1477,0.5009],
        [0.5148,0.2009,0.1027,0.1283,0.1236,0.1448,0.2274,0.2397,0.2871,0.2899,0.1721,0.2493],
        [0.5481,0.3237,0.0960,0.1200,0.0985,0.1458,0.1033,0.1344,0.1695,0.2797,0.0906,0.1265],
        [0.2738,0.6288,0.8501,0.1307,0.3452,0.1773,0.2324,0.1795,0.2856,0.5638,1.0651,0.3676],
        [0.2967,0.5518,1.3134,0.1427,0.2334,0.0876,0.1120,0.1279,0.2175,0.4896,0.1997,0.1785],
    ]
    _PREVIEW_PROT3 = [
        [0.0454,0.9775,0.9820,0.6197,0.4558,0.1757,0.2937,0.2065,0.2814,1.9233,0.9583,0.5525],
        [0.0473,1.0011,1.1122,0.3542,0.3131,0.2133,0.3345,0.1753,0.1665,0.4205,0.1922,0.4157],
        [0.0456,0.1673,0.9975,0.3380,0.7276,0.6222,0.2707,0.2433,0.3046,0.5361,0.7468,0.5336],
        [0.0472,0.1427,1.0385,0.3711,0.2572,0.8755,0.1648,0.1663,0.2339,0.0925,0.4129,0.2646],
        [0.5304,0.3151,0.5987,0.1605,0.2135,0.2375,0.6073,0.2831,1.1596,0.8248,0.5017,0.3201],
        [0.6298,0.3429,0.4897,0.1514,0.1690,0.2736,0.2881,0.1539,0.5136,0.5700,0.0448,0.1963],
        [0.1439,0.4296,2.5722,0.1251,0.3994,0.4783,0.0796,0.9855,0.2624,1.3059,0.6046,0.3306],
        [0.1637,0.5393,2.9379,0.1031,0.2104,0.2238,0.0598,0.6423,0.2201,0.6970,0.1377,0.1834],
    ]
    try:
        data = request.get_json(force=True)
        data['nProteins'] = 3
        data['nPlates']   = 1
        pnames = list(data.get('proteinNames') or [])
        while len(pnames) < 3:
            pnames.append(f'Protein {len(pnames) + 1}')
        data['proteinNames'] = pnames[:3]
        data['plates'] = [{
            'sampleGrid':   _PREVIEW_SAMPLE,
            'proteinGrids': [_PREVIEW_PROT1, _PREVIEW_PROT2, _PREVIEW_PROT3],
            'notes':        data.get('notes', ''),
        }]
        wb, date_str, _ = _elisa_run_generate(data)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name='ELISA_format_preview.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)