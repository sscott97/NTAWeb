import os
import json
import csv
import math
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from io import BytesIO

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

CONFIG_PATH = os.path.join(BASE_DIR, "config.json")
SETTINGS_PATH = os.path.join(BASE_DIR, "settings.json")

def load_config():
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, "r") as f:
            return json.load(f)
    return {}

config = load_config()

def load_csv_blocks(csv_stream):
    blocks = []
    current_block = []
    csv_stream.seek(0)
    reader = csv.reader(line.decode() for line in csv_stream.readlines())
    for row in reader:
        if not any(cell.strip() for cell in row):
            if current_block:
                blocks.append(current_block)
                current_block = []
        else:
            current_block.append(row[:12])
    if current_block:
        blocks.append(current_block)
    return blocks


def detect_csv_mode(csv_stream):
    """
    Examine a CSV to determine whether it looks like Standard or Data Only format.
    Returns 'standard', 'data_only', or 'unknown'.
    Does NOT consume the stream — resets to 0 before returning.
    """
    csv_stream.seek(0)
    raw = csv_stream.read()
    csv_stream.seek(0)

    if isinstance(raw, bytes):
        text = raw.decode("utf-8-sig", errors="replace")
    else:
        text = raw

    lines = text.splitlines()

    # Standard mode signature: look for rows starting with A–H labels
    row_letters_found = set()
    for line in lines:
        cells = [c.strip() for c in line.split(",")]
        if cells and cells[0] in ("A", "B", "C", "D", "E", "F", "G", "H"):
            # Check that subsequent cells are numeric
            numeric_count = 0
            for val in cells[1:13]:
                try:
                    float(val)
                    numeric_count += 1
                except (ValueError, TypeError):
                    pass
            if numeric_count >= 6:
                row_letters_found.add(cells[0])

    has_standard_markers = len(row_letters_found) >= 4  # at least A-D found

    # Data Only signature: look for rows that are pure numeric 12-column blocks
    pure_numeric_rows = 0
    for line in lines:
        cells = [c.strip() for c in line.split(",") if c.strip()]
        if len(cells) >= 12:
            try:
                [float(c) for c in cells[:12]]
                pure_numeric_rows += 1
            except (ValueError, TypeError):
                pass

    if has_standard_markers and pure_numeric_rows < 8:
        return "standard"
    elif pure_numeric_rows >= 8 and not has_standard_markers:
        return "data_only"
    elif has_standard_markers:
        return "standard"
    else:
        return "unknown"


def validate_csv_mode(csv_stream, selected_mode):
    """
    Check whether the selected CSV mode matches what the file actually looks like.
    Returns (ok: bool, detected_mode: str, message: str).
    """
    detected = detect_csv_mode(csv_stream)
    csv_stream.seek(0)

    if detected == "unknown":
        return True, detected, ""  # Can't determine — let it proceed

    if detected == selected_mode:
        return True, detected, ""

    if selected_mode == "standard" and detected == "data_only":
        return False, detected, (
            "This CSV appears to be in Data Only format (raw numeric blocks without "
            "A–H row labels), but you selected Standard mode. Would you like to "
            "switch to Data Only mode?"
        )
    elif selected_mode == "data_only" and detected == "standard":
        return False, detected, (
            "This CSV appears to be in Standard format (with A–H row labels and "
            "machine headers), but you selected Data Only mode. Would you like to "
            "switch to Standard mode?"
        )

    return True, detected, ""


def load_csv_blocks_standard(csv_stream):
    """
    Parse a Standard Mode CSV from the plate reader.

    In Standard Mode the 8×12 data blocks are bounded by row labels A–H
    down the left side and a header row with column numbers 1–12 above.
    This function locates every such block, strips the row letters and
    column header, and returns the same list-of-blocks structure as
    load_csv_blocks() so downstream processing is identical.
    """
    csv_stream.seek(0)
    raw_bytes = csv_stream.read()
    # Handle both bytes and str
    if isinstance(raw_bytes, bytes):
        text = raw_bytes.decode("utf-8-sig")
    else:
        text = raw_bytes

    lines = text.splitlines()

    ROW_LETTERS = [chr(65 + r) for r in range(8)]  # A–H

    blocks = []
    i = 0
    while i < len(lines):
        # Strip trailing commas that the instrument sometimes adds
        line = lines[i].strip().rstrip(",")
        cells = [c.strip() for c in line.split(",")]

        # Detect the start of a block: first cell is 'A' and at least 13
        # fields (letter + 12 data columns)
        if cells and cells[0] == "A" and len(cells) >= 13:
            block = []
            valid = True
            for r in range(8):
                row_idx = i + r
                if row_idx >= len(lines):
                    valid = False
                    break
                row_line = lines[row_idx].strip().rstrip(",")
                row_cells = [c.strip() for c in row_line.split(",")]
                if row_cells[0] != ROW_LETTERS[r]:
                    valid = False
                    break
                # Take columns 1–12 (skip the letter in column 0)
                block.append(row_cells[1:13])

            if valid and len(block) == 8:
                # Validate the block contains actual numeric data, not
                # placeholder dashes or empty cells.  At least one cell
                # in the 8×12 grid must be a valid number.
                has_numeric = False
                for row in block:
                    for val in row:
                        try:
                            float(val)
                            has_numeric = True
                            break
                        except (ValueError, TypeError):
                            pass
                    if has_numeric:
                        break
                if has_numeric:
                    # Require at least 12 numeric values (one full row's worth)
                    # to filter spurious summary/stats blocks the plate reader
                    # appends at the end of the CSV.
                    numeric_count = 0
                    for row in block:
                        for val in row:
                            try:
                                float(val)
                                numeric_count += 1
                            except (ValueError, TypeError):
                                pass
                    if numeric_count >= 12:
                        blocks.append(block)
            i += 8  # skip past this block regardless
        else:
            i += 1

    return blocks


def process_csv_to_template(
    csv_path,
    template_path,
    output_path,
    num_pseudotypes,
    pseudotype_texts,
    assay_title_text,
    sample_id_text,
    data_mode="data_only",
):
    if data_mode == "standard":
        blocks = load_csv_blocks_standard(csv_path)
    else:
        blocks = load_csv_blocks(csv_path)

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found at {template_path}")

    wb = openpyxl.load_workbook(template_path)
    template_sheet = wb.active

    pseudotype_list = [pt.strip() for line in pseudotype_texts.splitlines() for pt in line.split(",") if pt.strip()]

    # Pad with "Unlabelled" if not enough pseudotypes provided
    while len(pseudotype_list) < num_pseudotypes:
        pseudotype_list.append("Unlabelled")

    sample_id_list = [sid.strip() for line in sample_id_text.splitlines() for sid in line.split(",") if sid.strip()]

    sample_index = 0

    for i, block in enumerate(blocks):
        sheet_title = f"Plate{i+1}"
        new_sheet = wb.copy_worksheet(template_sheet)
        new_sheet.title = sheet_title

        for r in range(8):
            for c in range(12):
                cell = new_sheet.cell(row=5 + r, column=2 + c)
                try:
                    val = block[r][c]
                    cell.value = float(val) if val.replace('.', '', 1).isdigit() else val
                except IndexError:
                    cell.value = ""

        ws = new_sheet
        ws['B2'] = assay_title_text

        if num_pseudotypes == 1:
            val = pseudotype_list[0] if len(pseudotype_list) > 0 else ''
            for cell in ['B3', 'E3', 'H3', 'K3']:
                ws[cell] = val
            sample_cells = ['B4', 'E4', 'H4', 'K4']
            for cell in sample_cells:
                if sample_index < len(sample_id_list):
                    ws[cell] = sample_id_list[sample_index]
                    sample_index += 1
                else:
                    ws[cell] = ''
        elif num_pseudotypes == 2:
            ws['B3'] = pseudotype_list[0] if len(pseudotype_list) > 0 else ''
            ws['E3'] = pseudotype_list[0] if len(pseudotype_list) > 0 else ''
            ws['H3'] = pseudotype_list[1] if len(pseudotype_list) > 1 else ''
            ws['K3'] = pseudotype_list[1] if len(pseudotype_list) > 1 else ''
            val1 = sample_id_list[sample_index] if sample_index < len(sample_id_list) else ''
            if sample_index < len(sample_id_list): sample_index += 1
            val2 = sample_id_list[sample_index] if sample_index < len(sample_id_list) else ''
            if sample_index < len(sample_id_list): sample_index += 1
            ws['B4'] = val1
            ws['H4'] = val1
            ws['E4'] = val2
            ws['K4'] = val2
        elif num_pseudotypes == 3:
            ws['B3'] = pseudotype_list[0] if len(pseudotype_list) > 0 else ''
            ws['E3'] = pseudotype_list[1] if len(pseudotype_list) > 1 else ''
            ws['H3'] = pseudotype_list[2] if len(pseudotype_list) > 2 else ''
            ws['K3'] = ''
            val = sample_id_list[sample_index] if sample_index < len(sample_id_list) else ''
            if sample_index < len(sample_id_list): sample_index += 1
            ws['B4'] = val
            ws['E4'] = val
            ws['H4'] = val
            # Explicitly clear the unused 4th quadrant label and data
            ws['K4'] = ''
            for row in range(5, 13):
                for col in ['K', 'L', 'M']:
                    ws[f'{col}{row}'] = ''
        elif num_pseudotypes == 4:
            for idx, cell in enumerate(['B3', 'E3', 'H3', 'K3']):
                ws[cell] = pseudotype_list[idx] if idx < len(pseudotype_list) else ''
            val = sample_id_list[sample_index] if sample_index < len(sample_id_list) else ''
            if sample_index < len(sample_id_list): sample_index += 1
            for cell in ['B4', 'E4', 'H4', 'K4']:
                ws[cell] = val
        else:
            for cell in ['B3', 'E3', 'H3', 'K3', 'B4', 'E4', 'H4', 'K4']:
                ws[cell] = ''

    wb.remove(template_sheet)

    if isinstance(output_path, BytesIO):
        wb.save(output_path)
    else:
        wb.save(output_path)

def extract_final_titres_openpyxl(output_path):
    wb = load_workbook(output_path)

    if "Data Summary" in wb.sheetnames:
        wb.remove(wb["Data Summary"])

    summary_ws = wb.create_sheet("Data Summary", 0)

    summary_ws.append([
        "Plate", "Pseudotype", "Sample ID", 
        "NT 90% Replicate 1", "NT 90% Replicate 2", "NT 90% Replicate 3", "NT 90%",
        "NT 50% Replicate 1", "NT 50% Replicate 2", "NT 50% Replicate 3", "NT 50%"
    ])

    for col in range(1, 12):
        cell = summary_ws.cell(row=1, column=col)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    


    # Excel cell reference location 
    nt90_cells = [["B14", "C14", "D14"], ["E14", "F14", "G14"], ["H14", "I14", "J14"], ["K14", "L14", "M14"]]
    nt50_cells = [["B16", "C16", "D16"], ["E16", "F16", "G16"], ["H16", "I16", "J16"], ["K16", "L16", "M16"]]
    pseudotype_cells = ["B3", "E3", "H3", "K3"]
    sample_id_cells = ["B4", "E4", "H4", "K4"]
    nt90_avg_cells = ["C19", "F19", "I19", "L19"]
    nt50_avg_cells = ["C21", "F21", "I21", "L21"]

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Plate"):
            continue

        ws = wb[sheet_name]

        # Count how many pseudotypes are filled
        filled_pseudotypes = [
            bool(ws[cell].value and str(ws[cell].value).strip()) 
            for cell in pseudotype_cells
        ]
        num_pseudotypes = sum(filled_pseudotypes)

        # Only include as many columns as valid pseudotypes (up to 4, but can be 3 or fewer)
        for i in range(num_pseudotypes):
            pt_formula = f'=IF(TRIM({sheet_name}!{pseudotype_cells[i]})="", "Unlabelled", {sheet_name}!{pseudotype_cells[i]})'
            sid_cell = sample_id_cells[i]
            sid_formula = f'=IF(TRIM({sheet_name}!{sid_cell})="", "Unlabelled", {sheet_name}!{sid_cell})'
            nt90_formulas = [f"={sheet_name}!{cell}" for cell in nt90_cells[i]]
            nt90_avg = f"={sheet_name}!{nt90_avg_cells[i]}"
            nt50_formulas = [f"={sheet_name}!{cell}" for cell in nt50_cells[i]]
            nt50_avg = f"={sheet_name}!{nt50_avg_cells[i]}"

            summary_ws.append([
                sheet_name,
                pt_formula,
                sid_formula,
                *nt90_formulas,
                nt90_avg,
                *nt50_formulas,
                nt50_avg
            ])

            last_row = summary_ws.max_row
            for col in range(4, 12):
                cell = summary_ws.cell(row=last_row, column=col)
                cell.number_format = '0'

    # === Apply cell colouring ===
    light_green = PatternFill(start_color="AFE1AF", end_color="AFE1AF", fill_type="solid")
    dark_green  = PatternFill(start_color="89D289", end_color="89D289", fill_type="solid")
    light_blue  = PatternFill(start_color="7CAFB1", end_color="7CAFB1", fill_type="solid")
    dark_blue   = PatternFill(start_color="5F9EA0", end_color="5F9EA0", fill_type="solid")

    for row in summary_ws.iter_rows(min_row=2, max_row=summary_ws.max_row, min_col=4, max_col=11):
        for idx, cell in enumerate(row, start=4):
            if idx in (4, 5, 6):  # NT90 Replicates
                cell.fill = light_blue
            elif idx == 7:  # NT90 average
                cell.fill = dark_blue
            elif idx in (8, 9, 10):  # NT50 Replicates
                cell.fill = light_green
            elif idx == 11:  # NT50 average
                cell.fill = dark_green

    # === Style header row ===
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col in range(1, 12):
        cell = summary_ws.cell(row=1, column=col)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    
        # === Colour header row to match columns ===
    summary_ws["D1"].fill = light_blue
    summary_ws["E1"].fill = light_blue # NT50% replicates
    summary_ws["F1"].fill = light_blue
    summary_ws["G1"].fill = dark_blue

    summary_ws["H1"].fill = light_green
    summary_ws["I1"].fill = light_green
    summary_ws["J1"].fill = light_green # NT90% replicates
    summary_ws["K1"].fill = dark_green


    if isinstance(output_path, BytesIO):
        output_path.seek(0)
        wb.save(output_path)
    else:
        wb.save(output_path)

    add_default_to_final_titres(output_path)


def add_default_to_final_titres(output_path):
    wb = openpyxl.load_workbook(output_path)
    summary = wb["Data Summary"]
    plate1 = wb["Plate1"]

    # Get dilution series start and end values from Plate1 A5 to A11
    a5_val = plate1["A5"].value
    a11_val = plate1["A11"].value

    try:
        a5_val = round(float(a5_val))
    except:
        a5_val = ""

    try:
        a11_val = round(float(a11_val))
    except:
        a11_val = ""


    for row in summary.iter_rows(min_row=2, max_row=summary.max_row, min_col=4, max_col=10):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                # leave formulas intact
                continue
            if (cell.value in (None, "")) and a5_val:
                cell.value = f"≤{a5_val}"
            elif isinstance(cell.value, (float, int)) and a11_val and cell.value > a11_val:
                cell.value = f"≥{a11_val}"

    for row in summary.iter_rows(min_row=2, max_row=summary.max_row, min_col=1, max_col=11):
        for cell in row:
            try:
                if isinstance(cell.value, (float, int)):
                    cell.value = int(round(cell.value))
            except:
                pass
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, 12):
        summary.column_dimensions[get_column_letter(col)].width = 15

    if isinstance(output_path, BytesIO):
        output_path.seek(0)
        wb.save(output_path)
    else:
        wb.save(output_path)


# ════════════════════════════════════════════════════════════════
# Error Flagging — triplicate and titre replicate outlier detection
# ════════════════════════════════════════════════════════════════

def _is_outlier_in_triple(values, threshold_log2=1.0):
    """
    Given exactly 3 numeric values, check if any single value is an outlier.
    An outlier is defined as having a log2 fold difference > threshold from
    the median of the other two values.
    
    Returns a list of booleans (same length as values) indicating which are outliers,
    or None if there aren't enough valid values to assess.
    """
    nums = []
    for v in values:
        try:
            n = float(v)
            if n > 0:
                nums.append(n)
            else:
                nums.append(None)
        except (ValueError, TypeError):
            nums.append(None)
    
    valid_count = sum(1 for n in nums if n is not None)
    if valid_count < 2:
        return None  # Not enough data to assess
    
    outlier_flags = [False, False, False]
    
    if valid_count == 3:
        # Check each value against the median of the other two
        for i in range(3):
            others = [nums[j] for j in range(3) if j != i and nums[j] is not None]
            if len(others) == 2:
                median_others = (others[0] + others[1]) / 2
                if median_others > 0 and nums[i] is not None:
                    fold_diff = abs(math.log2(nums[i] / median_others))
                    if fold_diff > threshold_log2:
                        outlier_flags[i] = True
    elif valid_count == 2:
        # With only 2 values, check if they differ by more than threshold
        valid_vals = [(i, nums[i]) for i in range(3) if nums[i] is not None]
        if len(valid_vals) == 2:
            ratio = valid_vals[0][1] / valid_vals[1][1]
            if abs(math.log2(ratio)) > threshold_log2:
                # Flag both as potentially problematic
                outlier_flags[valid_vals[0][0]] = True
                outlier_flags[valid_vals[1][0]] = True
    
    if any(outlier_flags):
        return outlier_flags
    return None


def flag_triplicate_errors(output_path):
    """
    Scan all Plate sheets for triplicate outliers in raw luminescence data
    (rows 5-12) and in NT50/NT90 replicate values. Creates an 'Errors' sheet
    at the end of the workbook listing all flagged issues.
    
    Returns the number of errors found.
    """
    wb = openpyxl.load_workbook(output_path)
    
    # Remove existing Errors sheet if present
    if "Errors" in wb.sheetnames:
        wb.remove(wb["Errors"])
    
    errors_ws = wb.create_sheet("Errors")
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    section_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    section_font = Font(bold=True, color="922B21")
    warn_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    # Header row
    headers = [
        "Error Type", "Plate", "Quadrant", "Pseudotype", "Sample ID",
        "Dilution / Metric", "Rep 1", "Rep 2", "Rep 3",
        "Flagged Replicate(s)", "Log₂ Fold Diff"
    ]
    errors_ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = errors_ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # Column widths
    col_widths = [18, 10, 10, 18, 18, 18, 12, 12, 12, 20, 14]
    for i, w in enumerate(col_widths, 1):
        errors_ws.column_dimensions[get_column_letter(i)].width = w
    
    # Quadrant definitions
    quadrant_defs = [
        {'name': 'Q1', 'pt_cell': 'B3', 'sid_cell': 'B4', 'data_cols': ['B', 'C', 'D'],
         'nt90_cells': ['B14', 'C14', 'D14'], 'nt50_cells': ['B16', 'C16', 'D16']},
        {'name': 'Q2', 'pt_cell': 'E3', 'sid_cell': 'E4', 'data_cols': ['E', 'F', 'G'],
         'nt90_cells': ['E14', 'F14', 'G14'], 'nt50_cells': ['E16', 'F16', 'G16']},
        {'name': 'Q3', 'pt_cell': 'H3', 'sid_cell': 'H4', 'data_cols': ['H', 'I', 'J'],
         'nt90_cells': ['H14', 'I14', 'J14'], 'nt50_cells': ['H16', 'I16', 'J16']},
        {'name': 'Q4', 'pt_cell': 'K3', 'sid_cell': 'K4', 'data_cols': ['K', 'L', 'M'],
         'nt90_cells': ['K14', 'L14', 'M14'], 'nt50_cells': ['K16', 'L16', 'M16']},
    ]
    
    # Get dilution labels from first plate
    dilution_labels = []
    first_plate = None
    for sn in wb.sheetnames:
        if sn.startswith("Plate"):
            first_plate = sn
            break
    if first_plate:
        ws_fp = wb[first_plate]
        for row in range(5, 13):
            val = ws_fp[f'A{row}'].value
            try:
                num_val = float(val)
                if num_val == 0:
                    dilution_labels.append("NSC")
                elif num_val >= 1000:
                    dilution_labels.append(f"1:{int(num_val):,}")
                elif num_val == int(num_val):
                    dilution_labels.append(f"1:{int(num_val)}")
                else:
                    dilution_labels.append(str(num_val))
            except (ValueError, TypeError):
                dilution_labels.append(str(val) if val else f"Row {row}")
    
    error_count = 0
    
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Plate"):
            continue
        
        ws = wb[sheet_name]
        
        for quad in quadrant_defs:
            pt_val = ws[quad['pt_cell']].value
            if not pt_val or not str(pt_val).strip():
                continue  # Skip unused quadrants
            
            pseudotype = str(pt_val).strip()
            sid_val = ws[quad['sid_cell']].value
            sample_id = str(sid_val).strip() if sid_val and str(sid_val).strip() else "Unlabelled"
            
            # ── Check raw luminescence triplicates (rows 5-12) ──
            for row_idx, excel_row in enumerate(range(5, 13)):
                values = []
                for col in quad['data_cols']:
                    values.append(ws[f'{col}{excel_row}'].value)
                
                outliers = _is_outlier_in_triple(values)
                if outliers is not None:
                    # Calculate the max log2 fold diff for display
                    nums = []
                    for v in values:
                        try:
                            nums.append(float(v))
                        except (ValueError, TypeError):
                            nums.append(None)
                    
                    max_fold = 0
                    for i in range(3):
                        if outliers[i] and nums[i] is not None:
                            others = [nums[j] for j in range(3) if j != i and nums[j] is not None]
                            if others:
                                median_others = sum(others) / len(others)
                                if median_others > 0:
                                    fold = abs(math.log2(nums[i] / median_others))
                                    max_fold = max(max_fold, fold)
                    
                    flagged_reps = ", ".join([f"Rep {i+1}" for i in range(3) if outliers[i]])
                    dil_label = dilution_labels[row_idx] if row_idx < len(dilution_labels) else f"Row {excel_row}"
                    
                    # Format values for display
                    display_vals = []
                    for v in values:
                        try:
                            display_vals.append(round(float(v)))
                        except (ValueError, TypeError):
                            display_vals.append("—")
                    
                    errors_ws.append([
                        "Raw Triplicate",
                        sheet_name,
                        quad['name'],
                        pseudotype,
                        sample_id,
                        dil_label,
                        display_vals[0],
                        display_vals[1],
                        display_vals[2],
                        flagged_reps,
                        f"{max_fold:.2f}"
                    ])
                    error_count += 1
            
            # ── Check NT90 replicates ──
            nt90_values = [ws[cell].value for cell in quad['nt90_cells']]
            nt90_outliers = _is_outlier_in_triple(nt90_values)
            if nt90_outliers is not None:
                nums = []
                for v in nt90_values:
                    try:
                        nums.append(float(v))
                    except (ValueError, TypeError):
                        nums.append(None)
                
                max_fold = 0
                for i in range(3):
                    if nt90_outliers[i] and nums[i] is not None:
                        others = [nums[j] for j in range(3) if j != i and nums[j] is not None]
                        if others:
                            median_others = sum(others) / len(others)
                            if median_others > 0:
                                fold = abs(math.log2(nums[i] / median_others))
                                max_fold = max(max_fold, fold)
                
                flagged_reps = ", ".join([f"Rep {i+1}" for i in range(3) if nt90_outliers[i]])
                display_vals = []
                for v in nt90_values:
                    try:
                        display_vals.append(round(float(v)))
                    except (ValueError, TypeError):
                        display_vals.append("—")
                
                errors_ws.append([
                    "NT90 Replicate",
                    sheet_name,
                    quad['name'],
                    pseudotype,
                    sample_id,
                    "NT90",
                    display_vals[0],
                    display_vals[1],
                    display_vals[2],
                    flagged_reps,
                    f"{max_fold:.2f}"
                ])
                error_count += 1
            
            # ── Check NT50 replicates ──
            nt50_values = [ws[cell].value for cell in quad['nt50_cells']]
            nt50_outliers = _is_outlier_in_triple(nt50_values)
            if nt50_outliers is not None:
                nums = []
                for v in nt50_values:
                    try:
                        nums.append(float(v))
                    except (ValueError, TypeError):
                        nums.append(None)
                
                max_fold = 0
                for i in range(3):
                    if nt50_outliers[i] and nums[i] is not None:
                        others = [nums[j] for j in range(3) if j != i and nums[j] is not None]
                        if others:
                            median_others = sum(others) / len(others)
                            if median_others > 0:
                                fold = abs(math.log2(nums[i] / median_others))
                                max_fold = max(max_fold, fold)
                
                flagged_reps = ", ".join([f"Rep {i+1}" for i in range(3) if nt50_outliers[i]])
                display_vals = []
                for v in nt50_values:
                    try:
                        display_vals.append(round(float(v)))
                    except (ValueError, TypeError):
                        display_vals.append("—")
                
                errors_ws.append([
                    "NT50 Replicate",
                    sheet_name,
                    quad['name'],
                    pseudotype,
                    sample_id,
                    "NT50",
                    display_vals[0],
                    display_vals[1],
                    display_vals[2],
                    flagged_reps,
                    f"{max_fold:.2f}"
                ])
                error_count += 1
    
    # ── Style data rows ──
    for row in errors_ws.iter_rows(min_row=2, max_row=errors_ws.max_row, min_col=1, max_col=len(headers)):
        error_type = row[0].value
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border
        # Colour-code by error type
        if error_type == "Raw Triplicate":
            row[0].fill = warn_fill
        elif error_type in ("NT90 Replicate", "NT50 Replicate"):
            row[0].fill = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")
    
    # If no errors, add a message
    if error_count == 0:
        errors_ws.append(["No errors detected — all triplicates within acceptable range"])
        errors_ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        cell = errors_ws.cell(row=2, column=1)
        cell.font = Font(italic=True, color="28A745")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    if isinstance(output_path, BytesIO):
        output_path.seek(0)
        wb.save(output_path)
    else:
        wb.save(output_path)
    
    return error_count


def count_errors_from_workbook(file_bytes):
    """
    Count the number of error rows in the Errors sheet without modifying the file.
    Returns (error_count, has_errors_sheet).
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        if "Errors" not in wb.sheetnames:
            return 0, False
        
        ws = wb["Errors"]
        # Count data rows (skip header row 1)
        error_count = 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            val = row[0].value
            if val and val not in ("No errors detected — all triplicates within acceptable range",):
                error_count += 1
        
        return error_count, True
    except Exception:
        return 0, False


def save_template_path(path, config_file=CONFIG_PATH):
    config = load_config()
    config["template_path"] = path
    with open(config_file, "w") as f:
        json.dump(config, f, indent=4)

def load_template_path(config_file=CONFIG_PATH):
    config = load_config()
    template_path = config.get("template_path")
    if not template_path or not os.path.exists(template_path):
        raise FileNotFoundError("Saved template path not found or does not exist.")
    return template_path

DEFAULT_SETTINGS = {
    "timestamp_in_filename": False,
    "error_flagging": False,
    "presets": {
        "default": {
            "Q1": "#ff7e79",
            "Q2": "#ffd479",
            "Q3": "#009193",
            "Q4": "#d783ff"
        }
    },
    "selected_preset": "default"
}

def load_settings():
    if os.path.exists(SETTINGS_PATH):
        with open(SETTINGS_PATH, "r") as f:
            settings = json.load(f)
        # Ensure new keys exist with defaults
        if "error_flagging" not in settings:
            settings["error_flagging"] = False
        return settings
    return DEFAULT_SETTINGS.copy()

def save_settings(settings):
    with open(SETTINGS_PATH, "w") as f:
        json.dump(settings, f, indent=4)


def generate_sigmoid_csv(excel_path_or_bytes, output_csv_path):
    """
    Generate sigmoidData.csv from processed Excel workbook.
    """
    import math
    import tempfile
    
    if isinstance(excel_path_or_bytes, BytesIO):
        excel_path_or_bytes.seek(0)
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(excel_path_or_bytes.read())
            temp_path = tmp.name
        
        wb = load_workbook(temp_path)
        
        import os
        os.remove(temp_path)
    else:
        wb = load_workbook(excel_path_or_bytes)
    
    all_rows = []
    debug_info = []
    sample_counter = 1
    
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Plate"):
            continue
        
        ws = wb[sheet_name]
        debug_info.append(f"Processing sheet: {sheet_name}")
        
        dilutions = []
        for row in range(5, 12):
            val = ws[f'A{row}'].value
            try:
                dilutions.append(float(val))
            except (ValueError, TypeError):
                dilutions.append(None)
        
        debug_info.append(f"  Dilutions: {dilutions}")
        
        dilution_log2 = []
        for d in dilutions:
            if d and d > 0:
                dilution_log2.append(-math.log2(d))
            else:
                dilution_log2.append(None)
        
        debug_info.append(f"  DilutionLog2: {dilution_log2}")
        
        quadrants = [
            {'virus_cell': 'B3', 'sample_cell': 'B4', 'data_cols': ['B', 'C', 'D'], 'nsc_cols': ['B', 'C', 'D']},
            {'virus_cell': 'E3', 'sample_cell': 'E4', 'data_cols': ['E', 'F', 'G'], 'nsc_cols': ['E', 'F', 'G']},
            {'virus_cell': 'H3', 'sample_cell': 'H4', 'data_cols': ['H', 'I', 'J'], 'nsc_cols': ['H', 'I', 'J']},
            {'virus_cell': 'K3', 'sample_cell': 'K4', 'data_cols': ['K', 'L', 'M'], 'nsc_cols': ['K', 'L', 'M']},
        ]
        
        for quad_idx, quad in enumerate(quadrants):
            virus = ws[quad['virus_cell']].value
            sample = ws[quad['sample_cell']].value
            
            debug_info.append(f"  Quadrant {quad_idx+1}: Virus={virus}, Sample={sample}")
            
            virus_empty = not virus or not str(virus).strip()
            sample_empty = not sample or not str(sample).strip()
            if virus_empty and sample_empty:
                debug_info.append(f"    Skipped: Both virus and sample are empty (unused quadrant)")
                continue
            
            nsc_values = []
            for col in quad['nsc_cols']:
                cell_ref = f"{col}12"
                val = ws[cell_ref].value
                try:
                    nsc_values.append(float(val))
                except (ValueError, TypeError):
                    pass
            
            if not nsc_values:
                debug_info.append(f"    Skipped: No valid NSC values")
                continue
            
            nsc_mean = sum(nsc_values) / len(nsc_values)
            
            if nsc_mean == 0:
                debug_info.append(f"    Skipped: NSC mean is zero")
                continue
            
            has_data = False
            for row in range(5, 12):
                for col in quad['data_cols']:
                    cell_ref = f"{col}{row}"
                    val = ws[cell_ref].value
                    try:
                        float(val)
                        has_data = True
                        break
                    except (ValueError, TypeError):
                        pass
                if has_data:
                    break
            
            if not has_data:
                debug_info.append(f"    Skipped: No data in dilution rows")
                continue
            
            if not virus or str(virus).strip() == '':
                virus_str = 'Unlabelled'
            else:
                virus_str = str(virus).strip()
            
            if not sample or str(sample).strip() == '':
                sample_str = f'Unlabelled{sample_counter}'
                sample_counter += 1
            else:
                sample_str = str(sample).strip()
            
            debug_info.append(f"    Processing as: Virus={virus_str}, Sample={sample_str}")
            debug_info.append(f"    NSC mean: {nsc_mean}")
            
            quad_data_count = 0
            for i, row in enumerate(range(5, 12)):
                rep_values = []
                for col in quad['data_cols']:
                    cell_ref = f"{col}{row}"
                    val = ws[cell_ref].value
                    try:
                        rep_values.append(float(val))
                    except (ValueError, TypeError):
                        rep_values.append(None)

                valid_values = [v for v in rep_values if v is not None]
                if not valid_values:
                    continue

                triplicate_mean = sum(valid_values) / len(valid_values)
                neutralisation = 100 * (1 - (triplicate_mean / nsc_mean))

                if dilution_log2[i] is not None:
                    all_rows.append({
                        'Batch': 1,
                        'Virus': virus_str,
                        'Sample': sample_str,
                        'Dilution': dilutions[i],
                        'DilutionLog2': dilution_log2[i],
                        'Rep1': rep_values[0] if len(rep_values) > 0 and rep_values[0] is not None else '',
                        'Rep2': rep_values[1] if len(rep_values) > 1 and rep_values[1] is not None else '',
                        'Rep3': rep_values[2] if len(rep_values) > 2 and rep_values[2] is not None else '',
                        'Rep_Mean': round(triplicate_mean, 4),
                        'NSC_Mean': round(nsc_mean, 4),
                        'Neutralisation': round(neutralisation, 4)
                    })
                    quad_data_count += 1

            debug_info.append(f"    Added {quad_data_count} data points")

    if all_rows:
        with open(output_csv_path, 'w', newline='') as csvfile:
            fieldnames = ['Batch', 'Virus', 'Sample', 'Dilution', 'DilutionLog2', 'Rep1', 'Rep2', 'Rep3', 'Rep_Mean', 'NSC_Mean', 'Neutralisation']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(all_rows)
        print(f"Successfully generated sigmoid CSV with {len(all_rows)} data points")
    else:
        debug_str = "\n".join(debug_info)
        raise ValueError(f"No valid data found in Plate sheets to generate sigmoid CSV.\n\nDebug info:\n{debug_str}")
    
    return output_csv_path



def extract_nt50_titres_to_csv(excel_path, output_csv_path):
    """
    Directly reads the source values from Plate sheets (C21, F21, I21, L21)
    and saves them to a CSV so R doesn't have to deal with Excel formulas.
    """
    from openpyxl import load_workbook
    import csv

    wb = load_workbook(excel_path, data_only=True)
    
    nt50_avg_cells = ["C21", "F21", "I21", "L21"]
    pseudotype_cells = ["B3", "E3", "H3", "K3"]
    sample_id_cells = ["B4", "E4", "H4", "K4"]

    data_for_r = []

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Plate"):
            continue
        
        ws = wb[sheet_name]
        
        for i in range(4):
            pt_val = ws[pseudotype_cells[i]].value
            if pt_val and str(pt_val).strip():
                data_for_r.append({
                    'Pseudotype': str(pt_val).strip(),
                    'Sample_ID': str(ws[sample_id_cells[i]].value or "Unlabelled").strip(),
                    'NT50': ws[nt50_avg_cells[i]].value
                })

    with open(output_csv_path, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['Pseudotype', 'Sample_ID', 'NT50'])
        writer.writeheader()
        writer.writerows(data_for_r)
    
    return output_csv_path